"""ADR-0066 — style-only ghost of outside-block cells (xl3 0.8.1 sync).

Upstream (TS) bug, fixed in xl3 0.8.1 (commit 174bba1): the splice-based
restore pass moved outside-block cells back to their original rows after
a data-block expansion, but cleared only the VALUE at the shifted
position — borders/fills stayed behind. Large expansions rendered an
empty, fully-bordered ghost copy of the side summary block below the
data (production case: a 348-row settlement sheet left a 10-row ghost
at rows 352-361).

The Python port composes rows in place via an out-row cursor instead of
splice+restore (same model as the upstream wasm core, which was never
affected), so there is no shifted position to leak style from. These
tests pin the upstream regression property anyway, guarding the
invariant against a future change of emission strategy:

    no cell in the side columns may carry ink (border/fill)
    without a value.

Ported from upstream `renderer-outside-block-ghost-style.test.ts`.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, PatternFill, Side

from xl3 import convert

THIN = Side(style="thin")
BORDER = Border(top=THIN, left=THIN, bottom=THIN, right=THIN)
FILL = PatternFill(fill_type="solid", fgColor="FFDDEBF7")

# Side-summary columns used by both templates: P=16, Q=17.
SIDE_COLS = (16, 17)


def _has_ink(cell: Any) -> bool:
    """True when the cell carries any visible border or pattern fill."""
    b = cell.border
    has_border = b is not None and any(
        side is not None and side.style
        for side in (b.top, b.left, b.bottom, b.right)
    )
    f = cell.fill
    has_fill = (
        f is not None
        and getattr(f, "patternType", None) not in (None, "none")
    )
    return has_border or has_fill


def _is_empty_value(v: Any) -> bool:
    return v is None or v == ""


def _styled_empty_cells(ws: Any, cols: tuple[int, ...]) -> list[str]:
    """Collect addresses of cells in `cols` that have ink but no value."""
    ghosts: list[str] = []
    for r in range(1, (ws.max_row or 1) + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            if _has_ink(cell) and _is_empty_value(cell.value):
                ghosts.append(cell.coordinate)
    return ghosts


def _add_config(wb: Workbook) -> None:
    cfg = wb.create_sheet("__config__")
    cfg["A1"] = "key"
    cfg["B1"] = "value"
    cfg["A2"] = "source_sheet"
    cfg["B2"] = "Raw"
    cfg["A3"] = "source_table"
    cfg["B3"] = "1"


def _to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_data(n: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw"
    ws["A1"] = "a"
    ws["B1"] = "b"
    for i in range(n):
        ws.cell(row=i + 2, column=1).value = (
            "group-1" if i % 2 == 0 else "group-2"
        )
        ws.cell(row=i + 2, column=2).value = 100 * (i + 1)
    return _to_bytes(wb)


def _render(template: bytes, source: bytes) -> Any:
    outputs = convert(template, source)
    result = load_workbook(BytesIO(outputs[0].data))
    return result["Main"]


class TestOutsideBlockGhostStyle:
    def test_plain_path_side_summary_keeps_value_and_style_at_original_rows_only(
        self,
    ) -> None:
        """Plain render path: side summary below the block's template row."""
        wb = Workbook()
        wb.remove(wb.active)
        _add_config(wb)
        main = wb.create_sheet("Main")
        main["A3"] = "a"
        main["B3"] = "b"
        # Data block (col-scoped to A:B)
        main["A4"] = "{{ [a] }}"
        main["B4"] = "{{ [b] }}"
        # Side summary block BELOW the block's template row — in the
        # upstream splice impl these rows get shifted by the expansion
        # and must be restored in place without leaving a style ghost.
        for addr, value in (
            ("P5", "TOTAL"),
            ("Q5", 5500),
            ("P6", "TAX"),
            ("Q6", 550),
        ):
            cell = main[addr]
            cell.value = value
            cell.border = BORDER
            cell.fill = FILL

        sheet = _render(_to_bytes(wb), _make_data(10))

        # Preserved at original positions, with values and ink intact.
        assert sheet["P5"].value == "TOTAL"
        assert sheet["Q5"].value == 5500
        assert sheet["P6"].value == "TAX"
        assert sheet["Q6"].value == 550
        assert _has_ink(sheet["P5"])
        assert _has_ink(sheet["Q6"])

        # The upstream bug: shifted positions (row + insert count) kept
        # borders/fills after their values were cleared. No cell in the
        # side columns may carry ink without a value.
        assert _styled_empty_cells(sheet, SIDE_COLS) == []

    def test_grouped_path_side_cells_leave_no_ghost(self) -> None:
        """@group/@subtotal render path: side cell below the block."""
        wb = Workbook()
        wb.remove(wb.active)
        _add_config(wb)
        main = wb.create_sheet("Main")
        main["A1"] = "a"
        main["B1"] = "b"
        main["A2"] = "{{ @sort [a] }}"
        main["A3"] = "{{ @group [a] }}"
        main["A4"] = "{{ [a] }}"
        main["B4"] = "{{ [b] }}"
        main["A5"] = "Subtotal"
        main["B5"] = "{{ @subtotal SUM([b]) }}"
        # Side summary below the block (template rows 4-5).
        for addr, value in (("P6", "TOTAL"), ("Q6", 5500)):
            cell = main[addr]
            cell.value = value
            cell.border = BORDER
            cell.fill = FILL

        sheet = _render(_to_bytes(wb), _make_data(10))

        # The side cell survives exactly once, with value AND ink together…
        totals = 0
        for r in range(1, (sheet.max_row or 1) + 1):
            p = sheet.cell(row=r, column=16)
            if p.value == "TOTAL":
                totals += 1
                assert _has_ink(p)
                assert sheet.cell(row=r, column=17).value == 5500
        assert totals == 1

        # …and never as a style-only ghost.
        assert _styled_empty_cells(sheet, SIDE_COLS) == []
