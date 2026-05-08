"""Dynamic assertion evaluation (`runner-protocol.md` §"Dynamic assertions").

XTL 0.1 defines exactly one kind: `utc_today` — for each listed cell,
check that the rendered value matches the runner-start UTC date formatted
through XTL `TEXT()` date tokens.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .discover import DynamicCell


def _format_utc_today(now: datetime, fmt: str) -> str:
    """Render a UTC datetime using the XTL 0.1 `TEXT()` date tokens.

    Tokens (per language.md §"Text Formatting"):
      YYYY  4-digit year
      YY    2-digit year (right-most two digits)
      MM    zero-padded month
      DD/dd zero-padded day
      HH    zero-padded 24-hour
      hh    zero-padded 12-hour
      mm    zero-padded minute
      ss    zero-padded second

    For utc_today fixtures the format is typically YYYY-MM-DD; we still
    handle every token so a future fixture using e.g. YYYY/MM/DD works.
    """
    # Token replacement order matters: longer tokens first to avoid
    # collisions (YYYY before YY, etc.).
    out: list[str] = []
    i = 0
    while i < len(fmt):
        if fmt.startswith("YYYY", i):
            out.append(f"{now.year:04d}")
            i += 4
        elif fmt.startswith("YY", i):
            out.append(f"{now.year % 100:02d}")
            i += 2
        elif fmt.startswith("MM", i):
            out.append(f"{now.month:02d}")
            i += 2
        elif fmt.startswith("DD", i) or fmt.startswith("dd", i):
            out.append(f"{now.day:02d}")
            i += 2
        elif fmt.startswith("HH", i):
            out.append(f"{now.hour:02d}")
            i += 2
        elif fmt.startswith("hh", i):
            h12 = now.hour % 12 or 12
            out.append(f"{h12:02d}")
            i += 2
        elif fmt.startswith("mm", i):
            out.append(f"{now.minute:02d}")
            i += 2
        elif fmt.startswith("ss", i):
            out.append(f"{now.second:02d}")
            i += 2
        else:
            out.append(fmt[i])
            i += 1
    return "".join(out)


def check_utc_today(
    actual_files: list[tuple[str, bytes]],
    cells: list[DynamicCell],
    runner_start_utc: datetime,
) -> tuple[bool, str]:
    """Verify every listed dynamic cell matches the formatted runner-start
    UTC date. Returns (passed, diff)."""
    if not actual_files:
        return False, "no output files produced"

    # Per protocol, cells reference *the* output workbook by sheet/cell. The
    # fixture name pattern (`023-today-utc-dynamic`) implies a single file.
    # We look the cell up in the first output file. If a future fixture
    # needs multi-file dynamic addressing, the protocol will need a path
    # field on DynamicCell.
    _, data = actual_files[0]
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001
        return False, f"failed to load output workbook: {exc}"

    for cell in cells:
        if cell.sheet not in wb.sheetnames:
            return False, f"output workbook has no sheet {cell.sheet!r}"
        ws = wb[cell.sheet]
        actual: Any = ws[cell.cell].value
        expected = _format_utc_today(runner_start_utc, cell.format)
        # Compare as strings — the renderer should have written a string
        # for utc_today TEXT() output. Coerce numeric/date back to string
        # for diagnostic fidelity if the renderer wrote something else.
        if actual is None:
            return False, f"{cell.sheet}!{cell.cell}: expected {expected!r}, got <empty>"
        if str(actual) != expected:
            return False, f"{cell.sheet}!{cell.cell}: expected {expected!r}, got {actual!r}"
    return True, ""
