"""Stage 1 cell-value comparison via openpyxl.

Per `runner-protocol.md`:
> Stage 1 compares worksheet names and non-auxiliary cell values after loading
> .xlsx files through a spreadsheet library. This stage intentionally ignores
> styles, merges, page setup, embedded media, formulas beyond cached values,
> and package structure.

Stage 2 (canonical OOXML) is out of scope for the bootstrap port.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from ..value_model import canonical_string


def _cell_repr(c: Cell | Any) -> str:
    """Stringify a cell value for diff messages."""
    v = c.value if hasattr(c, "value") else c
    if v is None:
        return "<empty>"
    return repr(v)


def _values_equal(a: Any, b: Any) -> bool:
    """Compare two cell values for Stage 1 equality.

    Numbers compared by canonical-string form (so 1 and 1.0 both render as "1"
    and don't mismatch on a numeric round-trip; openpyxl returns ints for
    integral cell values when no decimal is stored). Everything else by Python
    equality, with empty (None / "") considered equal — Stage 1 cell-value
    comparison can't tell a blank cell from a cell whose value is "" because
    OOXML round-trips them identically.
    """
    if a is None and b is None:
        return True
    if a is None and b == "":
        return True
    if b is None and a == "":
        return True
    if a is None or b is None:
        return False
    # Numeric comparison via canonical form (handles 1 == 1.0)
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) and not isinstance(
        a, bool
    ) and not isinstance(b, bool):
        return canonical_string(float(a)) == canonical_string(float(b))
    return a == b


def _used_cells(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """Collect (row, col) → value for cells with a non-None value."""
    out: dict[tuple[int, int], Any] = {}
    if ws.max_row is None or ws.max_column is None:
        return out
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is not None and cell.value != "":
                out[(cell.row, cell.column)] = cell.value
    return out


def compare_workbooks(
    actual_bytes: bytes,
    expected_path: Path,
) -> tuple[bool, str]:
    """Stage 1 compare. Returns (passed, diff_message).

    diff_message is empty on pass; on failure, the first observed difference
    is returned in a stable form `<sheet>!<addr>: expected X, got Y`.
    """
    try:
        actual_wb = load_workbook(BytesIO(actual_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        return False, f"failed to load actual workbook: {exc}"
    try:
        expected_wb = load_workbook(expected_path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return False, f"failed to load expected workbook {expected_path}: {exc}"

    expected_sheet_names = list(expected_wb.sheetnames)
    actual_sheet_names = list(actual_wb.sheetnames)
    if expected_sheet_names != actual_sheet_names:
        return (
            False,
            f"sheet names differ: expected {expected_sheet_names!r}, got {actual_sheet_names!r}",
        )

    for sheet_name in expected_sheet_names:
        ews = expected_wb[sheet_name]
        aws = actual_wb[sheet_name]
        e_cells = _used_cells(ews)
        a_cells = _used_cells(aws)
        all_keys = sorted(set(e_cells) | set(a_cells))
        for key in all_keys:
            ev = e_cells.get(key)
            av = a_cells.get(key)
            if not _values_equal(ev, av):
                row, col = key
                from openpyxl.utils import get_column_letter

                addr = f"{get_column_letter(col)}{row}"
                return (
                    False,
                    f"{sheet_name}!{addr}: expected {ev!r}, got {av!r}",
                )
    return True, ""


def compare_workbook_dir(
    actual_files: list[tuple[str, bytes]],
    expected_dir: Path,
) -> tuple[bool, str]:
    """Compare a multi-file output (or zero-output) against `expected/`."""
    expected_files = sorted(p.name for p in expected_dir.iterdir() if p.suffix == ".xlsx")
    actual_filenames = sorted(name for name, _ in actual_files)
    if expected_files != actual_filenames:
        return (
            False,
            f"output filenames differ: expected {expected_files!r}, got {actual_filenames!r}",
        )
    actual_by_name = dict(actual_files)
    for fname in expected_files:
        passed, diff = compare_workbooks(actual_by_name[fname], expected_dir / fname)
        if not passed:
            return False, f"[{fname}] {diff}"
    return True, ""
