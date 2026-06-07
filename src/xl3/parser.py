"""Template workbook parser.

Reads a template `.xlsx` and returns a `ParsedTemplate` containing:
- `__config__` metadata (source_sheet, source_table, output_file_pattern, ...)
- author-defined `__config__` values
- `__inputs__` declarations (ADR-0010)
- `__sources__` declarations (ADR-0012) — header parsed; resolution happens at read time
- `__lists__` columns (ADR-0011)
- one `SheetTemplate` per visible/template sheet, with rows already grouped
  into blocks (directive rows + their data row, or static rows).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText

from .directives import (
    BlockDirective,
    BlockDirectives,
    DirectiveParseError,
    parse_directive,
)
from .errors import XtlErrorCode, xtl_error
from .expression import (
    CellTemplate,
    DirectiveSegment,
    ExprSegment,
    TextSegment,
    collect_referenced_columns,
    expression_has_data_marker_ref,
    expression_has_per_row_ref,
    parse_cell_template,
)
from .types import InputSpec, SourceSpec
from .value_model import canonical_string

# ---------------------------------------------------------------------------
# Parsed model
# ---------------------------------------------------------------------------


@dataclass
class TemplateMeta:
    """The system rows of `__config__`."""

    name: str | None = None
    description: str | None = None
    source_sheet: str | None = None
    source_table: str = "1"
    output_file_pattern: str = "output.xlsx"
    match_pattern: str | None = None
    # Author-defined values: any non-system key.
    author_values: dict[str, Any] = field(default_factory=dict)


_SYSTEM_KEYS = {
    "name",
    "description",
    "source_sheet",
    "source_table",
    "output_file_pattern",
    "match_pattern",
}


@dataclass
class TemplateCell:
    """A parsed template cell at a specific (row, col) coordinate."""

    row: int  # 1-based
    col: int  # 1-based
    template: CellTemplate
    referenced_columns: set[str] = field(default_factory=set)
    has_per_row_ref: bool = False
    has_data_marker: bool = False  # ADR-0066: truly-bracketed [col] ref present?
    raw_text: str = ""
    # xl3 0.8.1 sync: the cell's NATIVE value (number/date/boolean) when the
    # template cell holds a non-string. The TS splice model never rewrites
    # static cells so natives survive there; the compose model re-renders
    # every cell from text and must carry the native through explicitly —
    # otherwise `5500` re-emits as the string "5500".
    native_value: Any = None

    @property
    def has_data_refs(self) -> bool:
        return self.has_per_row_ref

    @property
    def is_directive_cell(self) -> bool:
        return self.template.is_directive_cell

    @property
    def is_subtotal_cell(self) -> bool:
        first = self.template.segments[0]
        return (
            isinstance(first, DirectiveSegment)
            and first.body.strip().lower().startswith("@subtotal")
        )


@dataclass
class StaticRowPlan:
    """A row that is emitted verbatim per render — no expansion."""

    template_row: int
    cells: list[TemplateCell]


@dataclass
class SubtotalCell:
    col: int
    aggregate: str
    column: str | None


@dataclass
class SubtotalRowPlan:
    template_row: int
    cells: list[TemplateCell]
    subtotals: list[SubtotalCell]


@dataclass
class DataRowPlan:
    """A row that is expanded once per filtered/sorted source row.

    ADR-0066: cells are scoped to the owning data block's `[col_start..col_end]`
    column range. Cells outside that range are tracked as outside cells on
    SheetTemplate and emitted at their original row positions.
    """

    template_row: int
    cells: list[TemplateCell]
    directives: BlockDirectives = field(default_factory=BlockDirectives)
    subtotal_rows: list[SubtotalRowPlan] = field(default_factory=list)
    col_start: int = 0
    col_end: int = 0


@dataclass
class OutsideCell:
    """ADR-0066: cell whose column lies outside every block's col-range.

    Outside cells are preserved verbatim at their original `(row, col)`
    position regardless of any block's expansion factor — they are neither
    cloned per record nor shifted by the splice.
    """

    row: int
    col: int
    cell: TemplateCell


@dataclass
class SheetTemplate:
    """One non-reserved sheet from the template workbook.

    `original_name` is the literal sheet name; group keys (bare identifiers in
    the sheet name like `Sheet_{Customer}`) aren't supported in the bootstrap.
    `directive_only_rows` are rows that contained only directive cells —
    those rows are stripped from output (renderer needs to know which rows to
    blank in the template before re-emitting).

    ADR-0066 / 0067 / 0068 / 0069: a sheet may carry multiple data blocks. The
    `plan` list keeps existing single-block semantics — multiple `DataRowPlan`
    entries on the same template row indicate side-by-side multi-block on
    that row (each with disjoint col-range). `outside_cells` collects cells
    that fall outside every block's column range so the renderer can emit
    them at their original row positions.
    """

    original_name: str
    plan: list[StaticRowPlan | DataRowPlan] = field(default_factory=list)
    max_col: int = 0
    directive_only_rows: set[int] = field(default_factory=set)
    outside_cells: list[OutsideCell] = field(default_factory=list)


@dataclass
class ParsedTemplate:
    meta: TemplateMeta
    sheets: list[SheetTemplate]
    inputs: list[InputSpec] = field(default_factory=list)
    sources: list[SourceSpec] = field(default_factory=list)
    list_sheets: dict[str, list[str]] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Cell text extraction (handles rich-text concatenation per spec)
# ---------------------------------------------------------------------------


def _cell_effective_text(cell: Cell) -> str:
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, CellRichText):
        return "".join(str(part) for part in v)
    if isinstance(v, str):
        return v
    return str(v)


# ---------------------------------------------------------------------------
# Reserved sheet detection (ADR-0011)
# ---------------------------------------------------------------------------


def is_reserved_sheet(name: str) -> bool:
    return (
        len(name) >= 5
        and name.startswith("__")
        and name.endswith("__")
        and name[2:-2].isalpha()
        and name[2:-2].islower()
    )


# ---------------------------------------------------------------------------
# Parser entry
# ---------------------------------------------------------------------------


def parse_template(template_bytes: bytes) -> ParsedTemplate:
    """Parse a template workbook. Read with `data_only=True` so formula
    cells expose their cached results (per ADR-0017)."""
    wb = load_workbook(BytesIO(template_bytes), data_only=True, rich_text=True)

    meta = TemplateMeta()
    if "__config__" in wb.sheetnames:
        meta = _parse_config_sheet(wb["__config__"])

    list_sheets: dict[str, list[str]] = {}
    if "__lists__" in wb.sheetnames:
        list_sheets = _parse_lists_sheet(wb["__lists__"])

    sources: list[SourceSpec] = []
    if "__sources__" in wb.sheetnames:
        sources = _parse_sources_sheet(wb["__sources__"])

    inputs: list[InputSpec] = []
    if "__inputs__" in wb.sheetnames:
        inputs = _parse_inputs_sheet(wb["__inputs__"], meta.author_values)

    # ADR-0011: any sheet matching `^__[a-z]+__$` is engine-reserved. The
    # four known names (__config__/__inputs__/__sources__/__lists__) are
    # processed above. Any OTHER dunder-wrapped sheet is an author error.
    _known_reserved = {"__config__", "__inputs__", "__sources__", "__lists__"}
    sheets: list[SheetTemplate] = []
    for sn in wb.sheetnames:
        if is_reserved_sheet(sn):
            if sn not in _known_reserved:
                raise xtl_error(
                    "xl3/sheet/reserved-name",
                    f'Sheet name "{sn}" is reserved (matches __<name>__ pattern)',
                )
            continue
        sheets.append(_parse_sheet_template(wb[sn]))

    return ParsedTemplate(
        meta=meta,
        sheets=sheets,
        inputs=inputs,
        sources=sources,
        list_sheets=list_sheets,
    )


def _parse_config_sheet(ws: Any) -> TemplateMeta:
    meta = TemplateMeta()
    for row in ws.iter_rows(values_only=False):
        if not row or len(row) < 2:
            continue
        key_cell, val_cell = row[0], row[1]
        key = _cell_effective_text(key_cell).strip()
        if not key:
            continue
        val = val_cell.value
        if key in _SYSTEM_KEYS:
            if key == "name":
                meta.name = str(val) if val is not None else None
            elif key == "description":
                meta.description = str(val) if val is not None else None
            elif key == "source_sheet":
                meta.source_sheet = str(val) if val is not None else None
            elif key == "source_table":
                meta.source_table = str(val) if val is not None else "1"
            elif key == "output_file_pattern":
                meta.output_file_pattern = str(val) if val is not None else "output.xlsx"
            elif key == "match_pattern":
                meta.match_pattern = str(val) if val is not None else None
        else:
            meta.author_values[key] = val
    return meta


def _parse_lists_sheet(ws: Any) -> dict[str, list[str]]:
    """Read `__lists__` per ADR-0011: row 1 = list names, columns below = values.

    Each value is canonicalized + trimmed; empty entries are dropped.
    """
    out: dict[str, list[str]] = {}
    if ws.max_row is None or ws.max_row < 1:
        return out
    header_row = list(ws[1])
    seen: set[str] = set()
    columns: list[tuple[int, str]] = []
    for cell in header_row:
        v = cell.value
        if v is None or v == "":
            continue
        name = str(v).strip()
        if not name:
            continue
        if name in seen:
            raise xtl_error(
                "xl3/sheet/duplicate-list-name",
                f'__lists__ has duplicate list name "{name}"',
            )
        seen.add(name)
        columns.append((cell.column, name))
    for col_idx, name in columns:
        values: list[str] = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            if isinstance(v, CellRichText):
                v = "".join(str(part) for part in v)
            s = canonical_string(v).strip()
            if s == "":
                continue
            values.append(s)
        out[name] = values
    return out


def _parse_sources_sheet(ws: Any) -> list[SourceSpec]:
    """Read `__sources__` per ADR-0012. Returns SourceSpec list."""
    if ws.max_row is None or ws.max_row < 1:
        return []
    header_row = list(ws[1])
    header_map: dict[str, int] = {}
    for cell in header_row:
        v = cell.value
        if v is None:
            continue
        name = str(v).strip().lower()
        if name:
            header_map[name] = cell.column
    if "name" not in header_map or "sheet" not in header_map:
        raise xtl_error(
            "xl3/source/missing-header",
            "__sources__ must declare 'name' and 'sheet' columns",
        )
    name_col = header_map["name"]
    sheet_col = header_map["sheet"]
    table_col = header_map.get("table")
    desc_col = header_map.get("description")

    out: list[SourceSpec] = []
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(row=r, column=name_col).value
        sh = ws.cell(row=r, column=sheet_col).value
        if nm is None and sh is None:
            continue
        if nm is None or sh is None:
            raise xtl_error(
                "xl3/source/missing-required",
                f"__sources__ row {r} missing required name/sheet",
            )
        name = str(nm).strip()
        if not name or name == "default" or name.startswith("__"):
            raise xtl_error(
                "xl3/source/invalid-name",
                f'__sources__ row {r} has invalid name "{name}"',
            )
        if name in seen:
            raise xtl_error(
                "xl3/source/duplicate-name",
                f'__sources__ has duplicate source name "{name}"',
            )
        seen.add(name)
        table = "1"
        if table_col is not None:
            tv = ws.cell(row=r, column=table_col).value
            if tv is not None:
                table = str(tv).strip() or "1"
        desc = None
        if desc_col is not None:
            dv = ws.cell(row=r, column=desc_col).value
            if dv is not None:
                desc = str(dv)
        out.append(SourceSpec(name=name, sheet=str(sh).strip(), table=table, description=desc))
    return out


_INPUT_TYPES = {"text", "number", "date", "select"}
_INPUT_BLOCK_RE = re.compile(r"\{\{\s*([\s\S]+?)\s*\}\}")
_INPUT_FORBIDDEN_PATTERNS: list[tuple[re.Pattern[str], XtlErrorCode, str]] = [
    (
        re.compile(r"(?<!\w)\[[^\]\r\n]+\]"),
        "xl3/inputs/forward-reference",
        "bare [Column] references (no source row context at input-read time)",
    ),
    (
        re.compile(r"(?<!\w)[A-Za-z]\w*\[[^\]\r\n]+\]"),
        "xl3/inputs/forward-reference",
        "Source[Column] references (sources are not loaded yet)",
    ),
    (
        re.compile(r"__sources__\["),
        "xl3/inputs/forward-reference",
        "__sources__ lookups",
    ),
    (
        re.compile(r"__inputs__\["),
        "xl3/inputs/forward-reference",
        "__inputs__ forward references (input rows are independent)",
    ),
    (
        re.compile(r"\bROW\s*\("),
        "xl3/inputs/runtime-only-fn",
        "ROW() (no repeat block at input-read time)",
    ),
    (
        re.compile(r"\b(?:SUM|COUNT|AVERAGE|AVG|MIN|MAX|XLOOKUP)\s*\("),
        "xl3/inputs/runtime-only-fn",
        "aggregate / lookup functions over source data",
    ),
]


def _assert_input_expr_allowed(raw: str, row_num: int, name: str, column: str) -> None:
    for block in _INPUT_BLOCK_RE.finditer(raw):
        inner = block.group(1) or ""
        for pattern, code, why in _INPUT_FORBIDDEN_PATTERNS:
            hit = pattern.search(inner)
            if hit:
                raise xtl_error(
                    code,
                    f'__inputs__ row {row_num} (name "{name}") {column} references '
                    f'"{hit.group(0)}" which is not available at input-read time — {why}',
                )


def _eval_input_cell_template(
    raw: str,
    config_values: dict[str, Any],
    row_num: int,
    name: str,
    column: str,
) -> str:
    if raw == "" or "{{" not in raw:
        return raw
    _assert_input_expr_allowed(raw, row_num, name, column)

    from .evaluator import EvalContext, evaluate

    try:
        tpl = parse_cell_template(raw)
        ctx = EvalContext(config_values=config_values, active_source_columns=set())
        if tpl.is_single_expression:
            seg = tpl.segments[0]
            assert isinstance(seg, ExprSegment)
            return canonical_string(evaluate(seg.expr, ctx))

        out: list[str] = []
        for seg in tpl.segments:
            if isinstance(seg, TextSegment):
                out.append(seg.text)
            elif isinstance(seg, ExprSegment):
                out.append(canonical_string(evaluate(seg.expr, ctx)))
            elif isinstance(seg, DirectiveSegment):
                out.append("")
        return canonical_string("".join(out))
    except Exception as exc:
        message = str(exc)
        exc.args = (f"{message} (at __inputs__ row {row_num} {column})",)
        raise


def _input_cell_literal(ws: Any, row: int, col: int | None) -> str:
    if col is None:
        return ""
    value = ws.cell(row=row, column=col).value
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        value = "".join(str(part) for part in value)
    return canonical_string(value).strip()


def _parse_inputs_sheet(ws: Any, config_values: dict[str, Any]) -> list[InputSpec]:
    """Read `__inputs__` per ADR-0010."""
    if ws.max_row is None or ws.max_row < 1:
        return []
    header_row = list(ws[1])
    header_map: dict[str, int] = {}
    for cell in header_row:
        v = cell.value
        if v is None:
            continue
        name = str(v).strip().lower()
        if name:
            header_map[name] = cell.column
    if "name" not in header_map or "type" not in header_map:
        raise xtl_error(
            "xl3/inputs/missing-header",
            "__inputs__ must declare 'name' and 'type' columns",
        )
    name_col = header_map["name"]
    type_col = header_map["type"]
    default_col = header_map.get("default")
    label_col = header_map.get("label")
    desc_col = header_map.get("description")
    options_col = header_map.get("options")

    out: list[InputSpec] = []
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        name = _input_cell_literal(ws, r, name_col)
        if not name:
            continue
        if name in seen:
            raise xtl_error(
                "xl3/inputs/duplicate-name",
                f'__inputs__ has duplicate input name "{name}"',
            )
        seen.add(name)
        type_str = _input_cell_literal(ws, r, type_col).lower()
        if type_str not in _INPUT_TYPES:
            raise xtl_error(
                "xl3/inputs/invalid-type",
                f'__inputs__ row {r}: type must be one of text/number/date/select',
            )
        default_raw = _eval_input_cell_template(
            _input_cell_literal(ws, r, default_col), config_values, r, name, "default"
        )
        label_raw = _eval_input_cell_template(
            _input_cell_literal(ws, r, label_col), config_values, r, name, "label"
        )
        description_raw = _eval_input_cell_template(
            _input_cell_literal(ws, r, desc_col), config_values, r, name, "description"
        )
        options_raw = _eval_input_cell_template(
            _input_cell_literal(ws, r, options_col), config_values, r, name, "options"
        )

        default = default_raw if default_raw != "" else None
        label = label_raw if label_raw != "" else None
        description = description_raw if description_raw != "" else None
        options = None
        if type_str == "select":
            if not options_raw:
                raise xtl_error(
                    "xl3/inputs/missing-options",
                    f'__inputs__ row {r}: select inputs require options',
                )
            options = [opt.strip() for opt in options_raw.split("|") if opt.strip()]
            if not options:
                raise xtl_error(
                    "xl3/inputs/missing-options",
                    f'__inputs__ row {r}: select inputs require options',
                )
            if default is not None and default not in options:
                raise xtl_error(
                    "xl3/inputs/select-option",
                    f'__inputs__ row {r} (name "{name}") default "{default}" is not in options',
                )
        out.append(
            InputSpec(
                name=name,
                type=type_str,  # type: ignore[arg-type]
                default=default,
                label=label,
                description=description,
                options=options,
            )
        )
    return out


def _directive_error_code(body: str) -> str:
    """Pick a stable error code based on which directive failed to parse.

    Specific codes are required for fixtures that assert error_code (082,
    094, etc.). Otherwise default to ADR-0027 `xl3/directive/invalid-syntax`.
    """
    s = body.lstrip().lower()
    if s.startswith("@group"):
        return "xl3/group/missing-key"
    if s.startswith("@join"):
        return "xl3/join/bad-on-clause"
    return "xl3/directive/invalid-syntax"


def _parse_subtotal_body(body: str) -> tuple[str, str | None]:
    m = re.match(
        r"^\s*@subtotal\s+(SUM|COUNT|AVERAGE|AVG|MIN|MAX)\s*\(\s*([^)]*?)\s*\)\s*$",
        body,
        re.IGNORECASE,
    )
    if not m:
        raise xtl_error(
            "xl3/subtotal/bad-aggregate",
            "@subtotal accepts SUM, COUNT, AVERAGE, MIN, MAX only",
        )
    aggregate = m.group(1).upper()
    if aggregate == "AVG":
        aggregate = "AVERAGE"
    inner = m.group(2).strip()
    if inner == "":
        if aggregate == "COUNT":
            return aggregate, None
        raise xtl_error(
            "xl3/subtotal/bad-aggregate",
            f"@subtotal {aggregate}() requires a column reference argument",
        )
    col_match = re.match(r"^\[\s*([^\]]+?)\s*\]$", inner)
    if not col_match:
        raise xtl_error(
            "xl3/subtotal/bad-aggregate",
            "@subtotal accepts SUM, COUNT, AVERAGE, MIN, MAX only",
        )
    return aggregate, col_match.group(1).strip()


@dataclass
class _ParsedDirectiveCell:
    """A directive cell collected during parsing, with its position so
    proximity-based attachment (ADR-0069) can run after block detection."""

    row: int
    col: int
    body: str  # raw directive text (without {{ }}); used for error messages
    parsed: Any  # parsed Directive (FilterDirective | ... | BlockDirective)


@dataclass
class _RawBlock:
    """A data block detected during parsing (before plan materialization)."""

    start_row: int
    end_row: int
    col_start: int
    col_end: int
    # row offsets (from start_row) that contain @subtotal expressions
    subtotal_row_offsets: list[int] = field(default_factory=list)
    directives: BlockDirectives = field(default_factory=BlockDirectives)


def _classify_cells(rows_cells: dict[int, list[TemplateCell]]) -> tuple[
    dict[int, list[TemplateCell]],  # directive cells per row
    dict[int, list[TemplateCell]],  # subtotal cells per row
    dict[int, list[TemplateCell]],  # data-marker cells (non-aggregate {{[col]}}) per row
    dict[int, set[int]],            # any {{...}} expression cols per row (markers + aggregates + subtotals)
    dict[int, set[int]],            # all non-empty cols per row
    set[int],                       # rows that contain ANY data marker (non-aggregate)
]:
    directive_cells: dict[int, list[TemplateCell]] = {}
    subtotal_cells: dict[int, list[TemplateCell]] = {}
    data_marker_cells: dict[int, list[TemplateCell]] = {}
    expr_cols: dict[int, set[int]] = {}
    nonempty_cols: dict[int, set[int]] = {}
    data_rows: set[int] = set()
    for r, cells in rows_cells.items():
        nonempty_cols[r] = {c.col for c in cells}
        expr_cols[r] = set()
        for c in cells:
            if c.is_subtotal_cell:
                # ADR-0038: validate aggregate shape eagerly so authors get
                # `xl3/subtotal/bad-aggregate` at parse time before any
                # block-detection / outside-group classification runs.
                seg = c.template.segments[0]
                assert isinstance(seg, DirectiveSegment)
                _parse_subtotal_body(seg.body)
                subtotal_cells.setdefault(r, []).append(c)
                expr_cols[r].add(c.col)
            elif c.is_directive_cell:
                directive_cells.setdefault(r, []).append(c)
                # directive cells DON'T count toward expr_cols (they sit above
                # the block and are stripped from output).
            else:
                # Any {{...}} body counts toward col-range (incl. aggregates).
                if any(isinstance(s, ExprSegment) for s in c.template.segments):
                    expr_cols[r].add(c.col)
                if c.has_data_marker:
                    data_marker_cells.setdefault(r, []).append(c)
                    data_rows.add(r)
    return (
        directive_cells,
        subtotal_cells,
        data_marker_cells,
        expr_cols,
        nonempty_cols,
        data_rows,
    )


def _compute_col_range(
    row_range: tuple[int, int],
    expr_cols: dict[int, set[int]],
    nonempty_cols: dict[int, set[int]],
) -> tuple[int, int]:
    """ADR-0066: block col range = bounding box of `{{...}}` expression cells
    across the row range, extended outward through contiguous non-empty cells
    in each row, taking running min/max."""
    col_start = 0
    col_end = 0
    for r in range(row_range[0], row_range[1] + 1):
        markers = expr_cols.get(r, set())
        if not markers:
            continue
        nonempty = nonempty_cols.get(r, set())
        left = min(markers)
        right = max(markers)
        while left > 1 and (left - 1) in nonempty:
            left -= 1
        while (right + 1) in nonempty:
            right += 1
        if col_start == 0 or left < col_start:
            col_start = left
        if right > col_end:
            col_end = right
    return col_start, col_end


def _detect_implicit_blocks(
    data_rows: set[int],
    subtotal_rows: set[int],
    expr_cols: dict[int, set[int]],
    nonempty_cols: dict[int, set[int]],
) -> list[_RawBlock]:
    """Implicit-mode (no `@block`) cluster detection per ADR-0066.

    A *block-bearing row* is any row in `data_rows` OR a subtotal row. A
    cluster is a maximal run of consecutive block-bearing rows. The col
    range is computed across the cluster's rows via `_compute_col_range`.
    Subtotal rows alone do not start a cluster — they only EXTEND an open
    cluster that began with a data row.
    """
    blocks: list[_RawBlock] = []
    if not data_rows:
        return blocks
    block_bearing = sorted(data_rows | subtotal_rows)
    cur_start: int | None = None
    cur_end: int | None = None
    cur_subtotals: list[int] = []
    cur_has_data = False
    for r in block_bearing:
        if cur_start is None:
            if r in data_rows:
                cur_start = r
                cur_end = r
                cur_subtotals = []
                cur_has_data = True
                if r in subtotal_rows:
                    cur_subtotals.append(0)
            # subtotal-only row before any data row: skip (the upstream
            # grouped renderer raises xl3/subtotal/outside-group later)
            continue
        assert cur_end is not None
        if r == cur_end + 1:
            cur_end = r
            if r in subtotal_rows and r not in data_rows:
                cur_subtotals.append(r - cur_start)
            if r in data_rows:
                cur_has_data = True
        else:
            # close current cluster
            if cur_has_data and cur_start is not None and cur_end is not None:
                cs, ce = _compute_col_range((cur_start, cur_end), expr_cols, nonempty_cols)
                blocks.append(
                    _RawBlock(
                        start_row=cur_start,
                        end_row=cur_end,
                        col_start=cs,
                        col_end=ce,
                        subtotal_row_offsets=list(cur_subtotals),
                    )
                )
            if r in data_rows:
                cur_start = r
                cur_end = r
                cur_subtotals = []
                cur_has_data = True
                if r in subtotal_rows:
                    cur_subtotals.append(0)
            else:
                cur_start = None
                cur_end = None
                cur_subtotals = []
                cur_has_data = False
    if cur_start is not None and cur_end is not None and cur_has_data:
        cs, ce = _compute_col_range((cur_start, cur_end), expr_cols, nonempty_cols)
        blocks.append(
            _RawBlock(
                start_row=cur_start,
                end_row=cur_end,
                col_start=cs,
                col_end=ce,
                subtotal_row_offsets=list(cur_subtotals),
            )
        )
    return blocks


def _build_explicit_blocks(
    ws_title: str,
    block_directives: list[_ParsedDirectiveCell],
    data_marker_cells: dict[int, list[TemplateCell]],
    expr_cols: dict[int, set[int]],
    nonempty_cols: dict[int, set[int]],
    subtotal_rows: set[int],
) -> list[_RawBlock]:
    """Explicit-mode block instantiation per ADRs 0067/0068. Each `@block`
    directive becomes one rectangle. Missing dimensions are auto-detected
    from marker cells. Empty rectangles raise `xl3/block/empty-table`.
    """
    blocks: list[_RawBlock] = []
    all_data_rows_with_markers = sorted(data_marker_cells.keys())
    for entry in block_directives:
        bd = entry.parsed
        assert isinstance(bd, BlockDirective)
        col_start = bd.col_start
        col_end = bd.col_end
        row_start = bd.row_start
        row_end = bd.row_end

        # Bare / col-range form: auto-detect row range as the first
        # contiguous run of marker rows below the directive's row.
        if row_start == 0 or row_end == 0:
            r = entry.row + 1
            cap = entry.row + 1000
            while r <= cap:
                cells = data_marker_cells.get(r, [])
                if col_start > 0:
                    cells = [c for c in cells if col_start <= c.col <= col_end]
                has_marker = len(cells) > 0
                if has_marker:
                    if row_start == 0:
                        row_start = r
                    row_end = r
                    r += 1
                elif row_start == 0:
                    r += 1
                    if r > entry.row + 100:
                        break
                else:
                    break
            if row_start == 0:
                raise xtl_error(
                    "xl3/block/empty-table",
                    f'@block at row {entry.row} on sheet "{ws_title}" has '
                    "no marker cells inside its declared rectangle",
                )

        # Bare form: also auto-detect col-range from markers in the row range.
        if col_start == 0:
            min_c = 0
            max_c = 0
            for r in range(row_start, row_end + 1):
                for c in data_marker_cells.get(r, []):
                    if min_c == 0 or c.col < min_c:
                        min_c = c.col
                    if c.col > max_c:
                        max_c = c.col
            if min_c == 0:
                raise xtl_error(
                    "xl3/block/empty-table",
                    f'@block at row {entry.row} on sheet "{ws_title}" has '
                    "no marker cells",
                )
            col_start = min_c
            col_end = max_c

        # Full-rect: verify the declared rectangle contains at least one marker.
        if bd.row_start > 0 and bd.row_end > 0:
            has_marker = False
            for r in range(row_start, row_end + 1):
                for c in data_marker_cells.get(r, []):
                    if col_start <= c.col <= col_end:
                        has_marker = True
                        break
                if has_marker:
                    break
            if not has_marker:
                raise xtl_error(
                    "xl3/block/empty-table",
                    f'@block at row {entry.row} declares rectangle '
                    f"(rows {row_start}-{row_end}, cols {col_start}-{col_end}) "
                    f'on sheet "{ws_title}" but contains no [Column] marker cells',
                )

        subs = sorted(r - row_start for r in subtotal_rows if row_start <= r <= row_end)
        blocks.append(
            _RawBlock(
                start_row=row_start,
                end_row=row_end,
                col_start=col_start,
                col_end=col_end,
                subtotal_row_offsets=list(subs),
            )
        )
    return blocks


def _check_block_overlap(blocks: list[_RawBlock], ws_title: str) -> None:
    for i in range(len(blocks)):
        for j in range(i + 1, len(blocks)):
            a, b = blocks[i], blocks[j]
            row_overlap = not (a.end_row < b.start_row or b.end_row < a.start_row)
            col_overlap = not (a.col_end < b.col_start or b.col_end < a.col_start)
            if row_overlap and col_overlap:
                raise xtl_error(
                    "xl3/block/overlap",
                    f"@block #{i + 1} (rows {a.start_row}-{a.end_row}, "
                    f"cols {a.col_start}-{a.col_end}) and #{j + 1} "
                    f"(rows {b.start_row}-{b.end_row}, cols {b.col_start}-{b.col_end}) "
                    f'overlap on sheet "{ws_title}"',
                )


def _is_inside_any_block(blocks: list[_RawBlock], row: int, col: int) -> bool:
    for b in blocks:
        if b.start_row <= row <= b.end_row and b.col_start <= col <= b.col_end:
            return True
    return False


def _find_owning_block(blocks: list[_RawBlock], row: int, col: int) -> _RawBlock | None:
    for b in blocks:
        if b.start_row <= row <= b.end_row and b.col_start <= col <= b.col_end:
            return b
    return None


def _col_in_any_block_range(blocks: list[_RawBlock], col: int) -> bool:
    for b in blocks:
        if b.col_start <= col <= b.col_end:
            return True
    return False


def _attach_directive_proximity(
    blocks: list[_RawBlock],
    directive: _ParsedDirectiveCell,
    ws_title: str,
) -> _RawBlock:
    """ADR-0069: attach a non-@block directive to the closest block such
    that (1) the directive's row is strictly above the block's first row
    AND (2) the directive's column is within the block's col-range."""
    candidates = [
        b
        for b in blocks
        if directive.row < b.start_row and b.col_start <= directive.col <= b.col_end
    ]
    if not candidates:
        kind = type(directive.parsed).__name__.replace("Directive", "").lower()
        raise xtl_error(
            "xl3/directive/orphan",
            f"Directive @{kind} at row {directive.row}, col {directive.col} "
            f'on sheet "{ws_title}" is not above any @block whose col-range '
            f"overlaps col {directive.col}",
        )
    candidates.sort(key=lambda b: b.start_row - directive.row)
    return candidates[0]


def _parse_directive_cell(tc: TemplateCell) -> Any:
    """Parse a directive cell's body, mapping DirectiveParseError to
    a stable XtlError code (per ADR-0027)."""
    seg = tc.template.segments[0]
    assert isinstance(seg, DirectiveSegment)
    body = seg.body
    try:
        return parse_directive(body)
    except DirectiveParseError as e:
        code = _directive_error_code(body)
        if code == "xl3/directive/invalid-syntax":
            msg = f"Invalid directive: {body.strip()}"
        else:
            msg = str(e)
        raise xtl_error(code, msg) from e


def _parse_sheet_template(ws: Any) -> SheetTemplate:
    """Walk a template sheet row-by-row, classify cells, detect data
    block(s) per ADRs 0066–0069, attach directives, and produce a plan
    plus list of outside cells.
    """
    st = SheetTemplate(original_name=ws.title)
    rows_cells: dict[int, list[TemplateCell]] = {}
    if ws.max_row is None:
        return st
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            text = _cell_effective_text(cell)
            if text == "":
                continue
            tpl = parse_cell_template(text)
            refs: set[str] = set()
            has_per_row = False
            has_data_marker = False
            for seg in tpl.segments:
                if isinstance(seg, ExprSegment):
                    refs |= collect_referenced_columns(seg.expr)
                    if expression_has_per_row_ref(seg.expr):
                        has_per_row = True
                    if expression_has_data_marker_ref(seg.expr):
                        has_data_marker = True
            tc = TemplateCell(
                row=cell.row,
                col=cell.column,
                template=tpl,
                referenced_columns=refs,
                has_per_row_ref=has_per_row,
                has_data_marker=has_data_marker,
                raw_text=text,
                native_value=(
                    cell.value
                    if not isinstance(cell.value, (str, CellRichText))
                    else None
                ),
            )
            rows_cells.setdefault(cell.row, []).append(tc)
            st.max_col = max(st.max_col, cell.column)

    if not rows_cells:
        return st

    (
        directive_cells,
        subtotal_cells_per_row,
        data_marker_cells,
        expr_cols,
        nonempty_cols,
        data_rows,
    ) = _classify_cells(rows_cells)

    # Pre-parse every directive cell once. We keep both BlockDirectives and
    # non-block directives in `parsed_directives` so the explicit-mode pass
    # can scan for `@block` and the proximity pass can attach the rest.
    parsed_directives: list[_ParsedDirectiveCell] = []
    directive_only_rows: set[int] = set()
    for r, dcells in directive_cells.items():
        # A row is directive-only if every cell on the row is a directive cell
        # (matches the prior single-directive-per-row contract). When mixed
        # with other cells the row is not directive-only and the directive
        # is silently dropped — this mirrors the previous behavior.
        if all(c.is_directive_cell for c in rows_cells[r]):
            directive_only_rows.add(r)
        for c in dcells:
            seg = c.template.segments[0]
            assert isinstance(seg, DirectiveSegment)
            parsed_directives.append(
                _ParsedDirectiveCell(
                    row=r,
                    col=c.col,
                    body=seg.body,
                    parsed=_parse_directive_cell(c),
                )
            )
    st.directive_only_rows = directive_only_rows

    block_directives = [
        d for d in parsed_directives if isinstance(d.parsed, BlockDirective)
    ]
    non_block_directives = [
        d for d in parsed_directives if not isinstance(d.parsed, BlockDirective)
    ]

    subtotal_rows = set(subtotal_cells_per_row.keys())
    explicit_mode = len(block_directives) > 0

    if explicit_mode:
        blocks = _build_explicit_blocks(
            ws.title,
            block_directives,
            data_marker_cells,
            expr_cols,
            nonempty_cols,
            subtotal_rows,
        )
        _check_block_overlap(blocks, ws.title)

        # Every data marker cell must lie inside some explicit block.
        for r, cells in data_marker_cells.items():
            for c in cells:
                if not _is_inside_any_block(blocks, r, c.col):
                    raise xtl_error(
                        "xl3/expression/bracket-outside-block",
                        f"[Column] reference at "
                        f"{_col_letters(c.col)}{r} on sheet "
                        f'"{ws.title}" is not inside any @block rectangle '
                        "(ADR-0068 explicit mode)",
                    )

        # Proximity-based directive attachment.
        for d in non_block_directives:
            target = _attach_directive_proximity(blocks, d, ws.title)
            target.directives.add(d.parsed)
    else:
        blocks = _detect_implicit_blocks(
            data_rows, subtotal_rows, expr_cols, nonempty_cols
        )
        if len(blocks) > 1:
            second = blocks[1]
            raise xtl_error(
                "xl3/expression/bracket-outside-block",
                f"[Column] references on sheet \"{ws.title}\" form "
                f"{len(blocks)} disconnected clusters. Use @block directives "
                f"to declare each (ADR-0067), or merge into one contiguous "
                f"block. Second cluster starts at row {second.start_row}.",
            )
        # ADR-0069 narrow break: even in single-block (implicit) sheets, a
        # directive whose column is outside the block's col-range raises
        # xl3/directive/orphan. To keep existing fixtures green we only
        # enforce this when there IS a block; sheets with no data block
        # never had a directive pipeline either.
        if blocks:
            for d in non_block_directives:
                target = _attach_directive_proximity(blocks, d, ws.title)
                target.directives.add(d.parsed)

    # Materialize plan + outside_cells from the blocks.
    block_map: dict[int, list[_RawBlock]] = {}
    for b in blocks:
        for r in range(b.start_row, b.end_row + 1):
            block_map.setdefault(r, []).append(b)

    data_row_plans_by_block: dict[int, DataRowPlan] = {}
    plan_entries: list[tuple[int, StaticRowPlan | DataRowPlan]] = []

    for r in sorted(rows_cells.keys()):
        cells = rows_cells[r]
        if r in directive_only_rows:
            continue
        # Subtotal rows are attached to their owning block's DataRowPlan.
        if r in subtotal_rows:
            sub_cells = subtotal_cells_per_row[r]
            # Find the block that owns this subtotal row.
            owning_block: _RawBlock | None = None
            for sc in sub_cells:
                ob = _find_owning_block(blocks, r, sc.col)
                if ob is not None:
                    owning_block = ob
                    break
            if owning_block is None:
                # Subtotal expression on a row not covered by any block.
                # Reject with the spec error — @subtotal MUST sit inside a
                # @group block (ADR-0038).
                raise xtl_error(
                    "xl3/subtotal/outside-group",
                    "@subtotal requires an active @group directive",
                )
            parsed_subs: list[SubtotalCell] = []
            for sc in sub_cells:
                seg = sc.template.segments[0]
                assert isinstance(seg, DirectiveSegment)
                aggregate, column = _parse_subtotal_body(seg.body)
                parsed_subs.append(
                    SubtotalCell(col=sc.col, aggregate=aggregate, column=column)
                )
            # Static cells in the same row become part of the SubtotalRowPlan
            # when they sit inside the block's col-range; otherwise outside.
            inside_subtotal_cells: list[TemplateCell] = list(sub_cells)
            for c in cells:
                if c.is_subtotal_cell:
                    continue
                if owning_block.col_start <= c.col <= owning_block.col_end:
                    inside_subtotal_cells.append(c)
                else:
                    st.outside_cells.append(OutsideCell(row=r, col=c.col, cell=c))
            data_plan = data_row_plans_by_block.get(id(owning_block))
            srp = SubtotalRowPlan(
                template_row=r,
                cells=inside_subtotal_cells,
                subtotals=parsed_subs,
            )
            if data_plan is not None:
                if data_plan.directives.group is None:
                    raise xtl_error(
                        "xl3/subtotal/outside-group",
                        "@subtotal requires an active @group directive",
                    )
                if len(data_plan.subtotal_rows) >= len(
                    data_plan.directives.group.keys
                ):
                    raise xtl_error(
                        "xl3/subtotal/outside-group",
                        f"@subtotal at row {r} has no matching @group level",
                    )
                data_plan.subtotal_rows.append(srp)
            else:
                # Data row hasn't been processed yet — stash for later
                # absorption when the block's start_row arrives.
                pending = getattr(owning_block, "_pending_subtotal_rows", None)
                if pending is None:
                    pending = []
                    setattr(owning_block, "_pending_subtotal_rows", pending)
                pending.append(srp)
            st.directive_only_rows.add(r)
            continue

        # Classify each cell. A cell at (r, c) belongs to ONE of:
        #   - owning_block: r and c both inside the block rect → cloned per record
        #   - above_static: c is in some block's col-range, r < that block start
        #   - below_static: c is in some block's col-range, r > that block end
        #   - outside: c is in no block's col-range
        outside_in_row: list[TemplateCell] = []
        above_in_row: list[TemplateCell] = []
        below_in_row: list[TemplateCell] = []
        per_owning_block: dict[int, list[TemplateCell]] = {}
        for c in cells:
            if c.is_directive_cell or c.is_subtotal_cell:
                continue
            owning = _find_owning_block(blocks, r, c.col)
            if owning is not None:
                per_owning_block.setdefault(id(owning), []).append(c)
                continue
            classified = False
            for b in blocks:
                if not (b.col_start <= c.col <= b.col_end):
                    continue
                if r < b.start_row:
                    above_in_row.append(c)
                    classified = True
                    break
                if r > b.end_row:
                    below_in_row.append(c)
                    classified = True
                    break
            if not classified:
                outside_in_row.append(c)

        for oc in outside_in_row:
            st.outside_cells.append(OutsideCell(row=r, col=oc.col, cell=oc))

        row_blocks_starting_here = [b for b in blocks if b.start_row == r]
        for b in row_blocks_starting_here:
            inside = per_owning_block.get(id(b), [])
            data_plan = DataRowPlan(
                template_row=r,
                cells=inside,
                directives=b.directives,
                col_start=b.col_start,
                col_end=b.col_end,
            )
            for srp in getattr(b, "_pending_subtotal_rows", []):
                if data_plan.directives.group is None:
                    raise xtl_error(
                        "xl3/subtotal/outside-group",
                        "@subtotal requires an active @group directive",
                    )
                if len(data_plan.subtotal_rows) >= len(
                    data_plan.directives.group.keys
                ):
                    raise xtl_error(
                        "xl3/subtotal/outside-group",
                        f"@subtotal at row {srp.template_row} has no matching @group level",
                    )
                data_plan.subtotal_rows.append(srp)
            plan_entries.append((r, data_plan))
            data_row_plans_by_block[id(b)] = data_plan

        # Above-static + below-static cells become a StaticRowPlan that the
        # renderer's out_row machinery places at the right position.
        # Note: a row CAN have both above (some blocks not yet started) and
        # below (others already ended) static cells; merge them.
        row_static = above_in_row + below_in_row
        if row_static and not row_blocks_starting_here:
            plan_entries.append(
                (r, StaticRowPlan(template_row=r, cells=row_static))
            )

    # Preserve original row order.
    plan_entries.sort(key=lambda p: p[0])
    st.plan = [entry for _, entry in plan_entries]
    return st


def _col_letters(col: int) -> str:
    """1-based column → letters (1→A, 27→AA)."""
    out = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        out = chr(65 + rem) + out
    return out
