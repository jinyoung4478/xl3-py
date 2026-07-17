"""Render output workbook(s) from parsed template + sources.

Operates on the block plan produced by `parser.py`:
  - StaticRowPlan: emit one row, expressions evaluated with no active row.
  - DataRowPlan:   apply directives → expand once per (filtered/sorted/top) source row.

Supports:
  - @source SourceName       (block iterates the named source)
  - @join JoinedSource on …  (inner-join: pair primary with first match)
  - @filter / @sort / @top   (transform the block's row set)
  - @repeat right [N]        (horizontal expansion)

NOT yet:
  - Multi-file groups (output_file_pattern with bare-ident group keys)
  - Sheet-name group keys
  - numFmt-driven coercion (ADR-0003)
  - Filename sanitization (ADR-0002)
"""

from __future__ import annotations

from copy import copy
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection

from .directives import (
    JoinDirective,
    apply_filters,
    apply_sorts,
    apply_top,
)
from .errors import xtl_error
from .evaluator import EvalContext, evaluate
from .expression import (
    CellTemplate,
    DirectiveSegment,
    ExprSegment,
    TextSegment,
)
from .grouper import group_key_canonical
from .parser import (
    DataRowPlan,
    OutsideCell,
    ParsedTemplate,
    SheetTemplate,
    StaticRowPlan,
    SubtotalCell,
    TemplateCell,
    is_reserved_sheet,
)
from .reader import SourceData
from .types import OutputFile
from .value_model import canonical_string, is_empty, parse_number_strict


def render(
    parsed: ParsedTemplate,
    sources: dict[str, SourceData],
    template_bytes: bytes,
    config_values: dict[str, Any] | None = None,
    inputs: dict[str, Any] | None = None,
) -> list[OutputFile]:
    """Render output files. ADR-0016 ordering: file groups by first-seen,
    sheet groups within a file by first-seen."""
    config_values = config_values or {}
    inputs = inputs or {}

    default = sources.get("default")
    default_rows = list(default.rows) if default else []

    file_group_keys = _extract_file_group_keys(parsed.meta.output_file_pattern or "")
    file_groups = _partition_first_seen(default_rows, file_group_keys)
    if not file_groups:
        # No source rows at all — emit a single file with empty data block
        # expansions (preserves headers / static content). Fixture 031 covers
        # the explicit zero-row case via `_should_suppress_output`.
        file_groups = [_GroupBucket(key=(), rows=[])]

    files: list[OutputFile] = []
    # ADR-0031: detect filename collisions across file groups before rendering
    # any of them. Two distinct group keys that sanitize to the same filename
    # would otherwise silently overwrite each other in host code.
    seen_filenames: set[str] = set()
    for bucket in file_groups:
        # Build a per-file sources dict with the default's rows narrowed to
        # this file's bucket. Named sources keep their full row sets.
        file_sources = dict(sources)
        if default is not None:
            file_sources = dict(sources)
            file_sources["default"] = SourceData(
                sheet_name=default.sheet_name,
                headers=list(default.headers),
                rows=list(bucket.rows),
            )

        wb = load_workbook(BytesIO(template_bytes))

        # Render each non-reserved sheet — splitting that sheet by sheet-group
        # keys when its name is a group-key template like `{{ Region }}`.
        for st in parsed.sheets:
            sheet_group_keys = _extract_sheet_group_keys(st.original_name)
            if sheet_group_keys:
                _render_grouped_sheet(
                    wb, st, sheet_group_keys, file_sources, config_values, inputs, parsed
                )
            else:
                ws = wb[st.original_name]
                _render_sheet(ws, st, file_sources, config_values, inputs, parsed)

        for sn in list(wb.sheetnames):
            if is_reserved_sheet(sn):
                del wb[sn]

        if _should_suppress_output(parsed, file_sources):
            continue

        out_io = BytesIO()
        wb.save(out_io)

        filename = _evaluate_filename(parsed, file_sources, config_values, inputs)
        sanitized, _warning = _sanitize_via_filename_module(filename)
        if sanitized in seen_filenames:
            from .errors import xtl_error

            raise xtl_error(
                "xl3/filename/collision",
                f'Output filename "{sanitized}" is produced by multiple file groups '
                "(their group key values collapse to the same sanitized filename). "
                "Make group keys distinct upstream — different cell values that "
                'share only forbidden characters (e.g., "Seoul/Korea" and "Seoul:Korea") '
                'both sanitize to "Seoul_Korea".',
            )
        seen_filenames.add(sanitized)
        files.append(OutputFile(filename=sanitized, data=out_io.getvalue()))
    return files


# ---------------------------------------------------------------------------
# Group-splitting helpers (ADR-0016)
# ---------------------------------------------------------------------------


from dataclasses import dataclass as _dc


@_dc
class _GroupBucket:
    key: tuple[Any, ...]
    rows: list[dict[str, Any]]


_BARE_IDENT_BLOCK_RE = __import__("re").compile(
    r"\{\{\s*([A-Za-z_][A-Za-z0-9_]*)\s*\}\}"
)
_BRACKET_REF_BLOCK_RE = __import__("re").compile(
    r"\{\{\s*\[\s*([^\]\r\n]+?)\s*\]\s*\}\}"
)


def _extract_group_keys(pattern: str) -> list[str]:
    """Both `{{ Region }}` and `{{ [Region] }}` in file/sheet patterns are
    group-key references per `language.md` §"Group Keys" + ADR-0026. They
    partition rows into file/sheet buckets. Expressions with operators,
    function calls, or `__config__[...]` / source-prefixed shapes are NOT
    group keys and evaluate against the bucket's first row instead.
    """
    keys: list[str] = []
    for m in _BARE_IDENT_BLOCK_RE.finditer(pattern):
        keys.append(m.group(1))
    for m in _BRACKET_REF_BLOCK_RE.finditer(pattern):
        keys.append(m.group(1))
    return keys


def _extract_file_group_keys(pattern: str) -> list[str]:
    return _extract_group_keys(pattern)


def _extract_sheet_group_keys(sheet_name: str) -> list[str]:
    return _extract_group_keys(sheet_name)


def _partition_first_seen(
    rows: list[dict[str, Any]], group_keys: list[str]
) -> list[_GroupBucket]:
    """Walk `rows` once; group by the tuple of `group_keys` values.

    Buckets are returned in **first-seen** order (per ADR-0016): the first
    row whose key is X causes bucket X to be emitted first.
    """
    if not group_keys:
        return [_GroupBucket(key=(), rows=list(rows))] if rows else []
    seen: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    order: list[tuple[Any, ...]] = []
    for r in rows:
        # ADR-0026: empty canonical-string values use the "(blank)" placeholder
        # so file/sheet names render predictably (e.g. "(blank).xlsx") rather
        # than producing a sanitization error or a generic fallback.
        key = tuple(group_key_canonical(r.get(k)) for k in group_keys)
        if key not in seen:
            seen[key] = []
            order.append(key)
        seen[key].append(r)
    return [_GroupBucket(key=k, rows=seen[k]) for k in order]


def _render_grouped_sheet(
    wb: Any,
    st: SheetTemplate,
    sheet_group_keys: list[str],
    sources: dict[str, SourceData],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
    parsed: ParsedTemplate,
) -> None:
    """Sheet-name template like `{{ Region }}`: clone the template sheet
    once per distinct sheet-group key combination (first-seen order) and
    render each with its bucket of rows.
    """
    default = sources.get("default")
    rows = list(default.rows) if default else []
    buckets = _partition_first_seen(rows, sheet_group_keys)
    if not buckets:
        # No source rows hit this file group — drop the template sheet entirely.
        if st.original_name in wb.sheetnames:
            del wb[st.original_name]
        return
    template_idx = wb.sheetnames.index(st.original_name)
    template_ws = wb[st.original_name]
    new_sheets: list[tuple[str, _GroupBucket]] = []
    for bucket in buckets:
        # Bucket keys are already canonical-string with the "(blank)"
        # placeholder applied by _partition_first_seen (ADR-0026), so the
        # resulting sheet name is always non-empty when there is at least
        # one group key.
        new_name = (
            str(bucket.key[0])
            if len(bucket.key) == 1
            else "_".join(str(v) for v in bucket.key)
        )
        if not new_name:
            new_name = st.original_name
        # Copy the template sheet using openpyxl's built-in copy
        copied = wb.copy_worksheet(template_ws)
        copied.title = new_name
        new_sheets.append((new_name, bucket))
    # Now remove the template sheet (it's been copied N times)
    del wb[st.original_name]
    # Move the copies to the template's original position
    for i, (name, bucket) in enumerate(new_sheets):
        ws = wb[name]
        # Restrict default source rows for this sheet's render
        per_sheet_sources = dict(sources)
        if default is not None:
            per_sheet_sources["default"] = SourceData(
                sheet_name=default.sheet_name,
                headers=list(default.headers),
                rows=list(bucket.rows),
            )
        _render_sheet(ws, st, per_sheet_sources, config_values, inputs, parsed)
        # Place the rendered sheet at the original index (keeps order tight)
        wb.move_sheet(ws, offset=template_idx + i - wb.sheetnames.index(name))


def _should_suppress_output(
    parsed: ParsedTemplate, sources: dict[str, SourceData]
) -> bool:
    """Fixture 031: when the default source has zero rows, produce no output.

    Per the ADR-0008 deferred gap, fixture 031 freezes the answer for the
    explicit zero-data `source_table` range case: no output workbook.
    """
    default = sources.get("default")
    return default is not None and not default.rows


def _sanitize_via_filename_module(filename: str) -> tuple[str, str | None]:
    from .filename import sanitize_filename

    return sanitize_filename(filename)


def _rename_group_sheets(
    wb: Any,
    parsed: ParsedTemplate,
    sources: dict[str, SourceData],
) -> None:
    """Apply `{{ <ident> }}` sheet-name templating using the first source row.

    Multi-group splitting (one rendered sheet per distinct group key) is the
    right answer per ADR-0016, but for the bootstrap we cover the single-row
    or single-group case which is what fixtures 086 needs.
    """
    import re as _re

    default_source = sources.get("default")
    if not default_source or not default_source.rows:
        return
    pattern = _re.compile(r"^\{\{\s*([A-Za-z_][A-Za-z0-9_]*)\s*\}\}$")
    for sn in list(wb.sheetnames):
        m = pattern.match(sn)
        if not m:
            continue
        col = m.group(1)
        first_val = default_source.rows[0].get(col)
        if first_val is None:
            continue
        new_name = canonical_string(first_val)
        if new_name and new_name != sn:
            wb[sn].title = new_name


def _render_sheet(
    ws: Any,
    st: SheetTemplate,
    sources: dict[str, SourceData],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
    parsed: ParsedTemplate,
) -> None:
    # Cache cell styles before mutating the worksheet.
    style_cache = _capture_styles(ws, st)
    # Capture styles for outside cells too so we can restore them.
    for oc in st.outside_cells:
        if (oc.row, oc.col) in style_cache:
            continue
        cell = ws.cell(row=oc.row, column=oc.col)
        style_cache[(oc.row, oc.col)] = (
            copy(cell.font),
            copy(cell.fill),
            copy(cell.border),
            copy(cell.alignment),
            copy(cell.number_format),
        )

    # Compute the row range we need to clear: all rows referenced by the plan
    # PLUS any directive-only rows PLUS rows of outside cells (so the original
    # template text gets blanked before we re-emit outside cells with their
    # parsed template values).
    template_rows_used: set[int] = set()
    subtotal_template_rows: set[int] = set()
    for plan in st.plan:
        template_rows_used.add(plan.template_row)
        if isinstance(plan, DataRowPlan):
            for srp in plan.subtotal_rows:
                template_rows_used.add(srp.template_row)
                subtotal_template_rows.add(srp.template_row)
    template_rows_used |= st.directive_only_rows
    template_rows_used |= {oc.row for oc in st.outside_cells}
    if template_rows_used:
        min_r = min(template_rows_used)
        max_r = ws.max_row or max(template_rows_used)
        for r in range(min_r, max_r + 1):
            for c in range(1, st.max_col + 1):
                ws.cell(row=r, column=c).value = None

    # Group plan entries by template_row so multi-block sheets emit all
    # blocks anchored at the same template row in parallel (each block at
    # its own col-range, sharing the same out_row anchor).
    grouped: list[tuple[int, list[StaticRowPlan | DataRowPlan]]] = []
    for plan in st.plan:
        if grouped and grouped[-1][0] == plan.template_row:
            grouped[-1][1].append(plan)
        else:
            grouped.append((plan.template_row, [plan]))

    out_row = min(template_rows_used) if template_rows_used else 1
    last_template_row: int | None = None
    for template_row, entries in grouped:
        if last_template_row is not None:
            gap = template_row - last_template_row - 1
            directive_rows_in_gap = sum(
                1
                for r in st.directive_only_rows
                if last_template_row < r < template_row
            )
            preserved_empty = gap - directive_rows_in_gap
            if preserved_empty > 0:
                out_row += preserved_empty

        # All entries on a shared template_row share the same out_row anchor.
        next_out_rows: list[int] = []
        for plan in entries:
            if isinstance(plan, StaticRowPlan):
                _emit_static(
                    ws, plan, out_row, style_cache, sources, config_values, inputs
                )
                next_out_rows.append(out_row + 1)
            else:
                after = _emit_data_block(
                    ws, plan, out_row, sources, style_cache, config_values, inputs
                )
                next_out_rows.append(after)

        out_row = max(next_out_rows) if next_out_rows else out_row
        # `last_template_row` advances to cover any subtotal rows in this group.
        last_template_row = template_row
        for plan in entries:
            if isinstance(plan, DataRowPlan):
                end = plan.template_row + len(plan.subtotal_rows)
                if end > last_template_row:
                    last_template_row = end

    # ADR-0066: outside cells preserved at their ORIGINAL row positions,
    # regardless of any block expansion. Emit AFTER the main plan so we
    # overwrite whatever (cleared/shifted) state lives at those coordinates.
    if st.outside_cells:
        _emit_outside_cells(
            ws,
            st.outside_cells,
            st.directive_only_rows - subtotal_template_rows,
            style_cache,
            sources,
            config_values,
            inputs,
        )


def _capture_styles(
    ws: Any, st: SheetTemplate
) -> dict[tuple[int, int], Any]:
    cache: dict[tuple[int, int], Any] = {}
    for plan in st.plan:
        rows = [plan.template_row]
        if isinstance(plan, DataRowPlan):
            rows.extend(srow.template_row for srow in plan.subtotal_rows)
        for row_num in rows:
            for c in range(1, st.max_col + 1):
                cell = ws.cell(row=row_num, column=c)
                cache[(row_num, c)] = (
                    copy(cell.font),
                    copy(cell.fill),
                    copy(cell.border),
                    copy(cell.alignment),
                    copy(cell.number_format),
                )
    return cache


def _apply_style(cell: Cell, style: Any) -> None:
    if style is None:
        return
    font, fill, border, align, fmt = style
    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = align
    cell.number_format = fmt


def _clear_cell_value_and_style(cell: Cell) -> None:
    cell.value = None
    cell.hyperlink = None
    cell.comment = None
    cell.font = Font()
    cell.fill = PatternFill()
    cell.border = Border()
    cell.alignment = Alignment()
    cell.protection = Protection()
    cell.number_format = "General"


def _emit_outside_cells(
    ws: Any,
    outside_cells: list[OutsideCell],
    directive_only_rows: set[int],
    style_cache: dict[tuple[int, int], Any],
    sources: dict[str, SourceData],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
) -> None:
    """ADR-0066: emit outside-block cells at their ORIGINAL `(row, col)`
    template positions. Outside cells are evaluated with the file/sheet
    group's first row as active row (so bare-identifier group-key refs
    still resolve) plus the usual cross-source aggregate machinery."""
    default_source = sources.get("default")
    active_row: dict[str, Any] = {}
    if default_source and default_source.rows:
        active_row = dict(default_source.rows[0])
    ctx = EvalContext(
        active_row=active_row,
        inputs=inputs,
        config_values=config_values,
        active_source_columns=None,
        named_sources=_build_named_sources_view(sources),
        active_row_set=list(default_source.rows) if default_source else None,
    )
    target_rows = {
        (oc.row, oc.col): oc.row - sum(1 for r in directive_only_rows if r < oc.row)
        for oc in outside_cells
    }
    for oc in outside_cells:
        target_row = target_rows[(oc.row, oc.col)]
        if target_row != oc.row:
            _clear_cell_value_and_style(ws.cell(row=oc.row, column=oc.col))

    for oc in outside_cells:
        target_row = target_rows[(oc.row, oc.col)]
        value = _render_cell(oc.cell, ctx)
        style = style_cache.get((oc.row, oc.col))
        if oc.cell.template.is_single_expression and style is not None:
            value = _apply_numfmt_coercion(value, style[4])
        target = _write_cell_value(ws, target_row, oc.col, value)
        _apply_style(target, style)


def _emit_static(
    ws: Any,
    plan: StaticRowPlan,
    out_row: int,
    style_cache: dict[tuple[int, int], Any],
    sources: dict[str, SourceData],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
) -> None:
    # Static cells can still reference cross-source aggregates and XLOOKUP
    # over named sources, so expose them on the context. The default
    # source's full row set serves as `active_row_set` for bare-bracket
    # aggregates that appear in static cells (e.g., a totals row).
    # ADR-0026 group keys: bare-identifier refs in a static cell resolve
    # against the file/sheet bucket's first row (the TS impl overlays the
    # group-key values onto the static context). Using the first row of the
    # filtered default source preserves the resolution for fixtures 006 /
    # 007 / 049 / 085 even though the cell is no longer classified as a
    # data marker row under ADR-0066.
    default_source = sources.get("default")
    active_row: dict[str, Any] = {}
    if default_source and default_source.rows:
        active_row = dict(default_source.rows[0])
    ctx = EvalContext(
        active_row=active_row,
        inputs=inputs,
        config_values=config_values,
        active_source_columns=None,
        named_sources=_build_named_sources_view(sources),
        active_row_set=list(default_source.rows) if default_source else None,
    )
    for tc in plan.cells:
        value = _render_cell(tc, ctx)
        style = style_cache.get((plan.template_row, tc.col))
        if tc.template.is_single_expression and style is not None:
            value = _apply_numfmt_coercion(value, style[4])
        target = _write_cell_value(ws, out_row, tc.col, value)
        _apply_style(target, style)


def _emit_data_block(
    ws: Any,
    plan: DataRowPlan,
    out_row: int,
    sources: dict[str, SourceData],
    style_cache: dict[tuple[int, int], Any],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
) -> int:
    """Emit one expanded data block; return the next free output row."""
    bd = plan.directives

    # Resolve the active source.
    active_source_name = bd.source_directive.source_name if bd.source_directive else "default"
    if active_source_name not in sources:
        raise xtl_error(
            "xl3/source/undeclared",
            f'Source "{active_source_name}" is not declared in __sources__',
        )
    primary = sources[active_source_name]

    # Apply filter / sort / top to the primary row set.
    rows = list(primary.rows)
    rows = apply_filters(rows, bd.filters, _collect_lists(plan, sources, config_values, inputs))
    rows = apply_sorts(rows, bd.sorts)
    rows = apply_top(rows, bd.top)

    # Resolve the join, if any.
    join: JoinDirective | None = bd.join_directive
    joined_rows_for_primary: list[dict[str, Any] | None] = [None] * len(rows)
    if join is not None:
        if join.joined_source not in sources:
            raise xtl_error(
                "xl3/join/undeclared-source",
                f'@join source "{join.joined_source}" must be declared in __sources__',
            )
        # The "other side" of the on-clause must name the block's active
        # source — otherwise the join can't be satisfied.
        if join.primary_source != active_source_name:
            raise xtl_error(
                "xl3/join/bad-on-clause",
                "@join key columns must reference the joined and primary sources",
            )
        joined_data = sources[join.joined_source]
        # Validate columns referenced by the on-clause exist on each side.
        if join.primary_column not in primary.headers:
            raise xtl_error(
                "xl3/source/unknown-column",
                f'Column "{join.primary_column}" does not exist in source "{join.primary_source}"',
            )
        if join.joined_column not in joined_data.headers:
            raise xtl_error(
                "xl3/source/unknown-column",
                f'Column "{join.joined_column}" does not exist in source "{join.joined_source}"',
            )
        kept: list[dict[str, Any]] = []
        kept_pairs: list[dict[str, Any]] = []
        for r in rows:
            primary_key = r.get(join.primary_column)
            match = _first_match(joined_data.rows, join.joined_column, primary_key)
            if match is None:
                continue  # inner-join semantics drop unmatched rows
            kept.append(r)
            kept_pairs.append(match)
        rows = kept
        joined_rows_for_primary = list(kept_pairs)  # type: ignore[assignment]

    if bd.group is not None:
        return _emit_grouped_block(
            ws,
            plan,
            out_row,
            rows,
            joined_rows_for_primary,
            join,
            sources,
            active_source_name,
            primary,
            style_cache,
            config_values,
            inputs,
        )

    # Repeat-right vs default vertical expansion.
    if bd.repeat_right is not None:
        return _emit_repeat_right(
            ws,
            plan,
            out_row,
            rows,
            joined_rows_for_primary,
            join,
            sources,
            active_source_name,
            primary,
            style_cache,
            config_values,
            inputs,
            bd.repeat_right.col_span,
        )
    return _emit_vertical(
        ws,
        plan,
        out_row,
        rows,
        joined_rows_for_primary,
        join,
        sources,
        active_source_name,
        primary,
        style_cache,
        config_values,
        inputs,
    )


def _collect_lists(
    plan: DataRowPlan,
    sources: dict[str, SourceData],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
) -> dict[str, list[str]]:
    # Lists live on the parsed template, not on per-block data; resolve from
    # the parser's result via a closure when we wire it in. Caller passes the
    # template in render(), but apply_filters needs lookup_lists at runtime.
    # We instead expose `lists` via plan? Cleaner: expose via a module-level
    # holder. Use a thread-local-style fallback since each render() call is
    # synchronous.
    return _RENDER_LISTS.get() or {}


# A simple holder so _emit_data_block/_collect_lists can see the parsed
# template's list_sheets without threading the dict through every call.
class _ListHolder:
    _stack: list[dict[str, list[str]]] = []

    def push(self, lists: dict[str, list[str]]) -> None:
        self._stack.append(lists)

    def pop(self) -> None:
        self._stack.pop()

    def get(self) -> dict[str, list[str]] | None:
        return self._stack[-1] if self._stack else None


_RENDER_LISTS = _ListHolder()


def _first_match(
    joined_rows: list[dict[str, Any]],
    joined_column: str,
    primary_key: Any,
) -> dict[str, Any] | None:
    from .value_model import compare_values

    for r in joined_rows:
        if compare_values(r.get(joined_column), primary_key) == 0:
            return r
    return None


def _numeric_values(rows: list[dict[str, Any]], column: str | None) -> list[float]:
    nums: list[float] = []
    if column is None:
        return nums
    for row in rows:
        value = row.get(column)
        if is_empty(value):
            continue
        if isinstance(value, bool):
            nums.append(1.0 if value else 0.0)
        elif isinstance(value, (int, float)):
            nums.append(float(value))
        else:
            parsed = parse_number_strict(value)
            if parsed is not None:
                nums.append(parsed)
    return nums


def _eval_subtotal(sc: SubtotalCell, rows: list[dict[str, Any]]) -> Any:
    if sc.aggregate == "COUNT":
        if sc.column is None:
            return len(rows)
        return sum(1 for row in rows if not is_empty(row.get(sc.column)))
    if sc.aggregate in ("SUM", "AVERAGE"):
        nums = _numeric_values(rows, sc.column)
        if sc.aggregate == "SUM":
            return sum(nums) if nums else 0
        return sum(nums) / len(nums) if nums else 0
    if sc.aggregate in ("MIN", "MAX") and sc.column is not None:
        from .evaluator import _aggregate_extremum

        return _aggregate_extremum(
            [row.get(sc.column) for row in rows], sc.aggregate.lower()
        )
    raise xtl_error(
        "xl3/subtotal/bad-aggregate",
        "@subtotal accepts SUM, COUNT, AVERAGE, MIN, MAX only",
    )


def _emit_grouped_block(
    ws: Any,
    plan: DataRowPlan,
    out_row: int,
    rows: list[dict[str, Any]],
    joined_rows: list[dict[str, Any] | None],
    join: JoinDirective | None,
    sources: dict[str, SourceData],
    active_source_name: str,
    primary: SourceData,
    style_cache: dict[tuple[int, int], Any],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
) -> int:
    from .grouper import partition_by_group_keys, plan_emission_events

    group = plan.directives.group
    if group is None:
        return out_row
    tree = partition_by_group_keys(rows, group.keys)
    events = plan_emission_events(tree, len(group.keys))
    joined_by_row = {id(row): joined_rows[i] for i, row in enumerate(rows)}
    data_index = 0

    for ev in events:
        if ev.kind == "data":
            assert ev.row is not None
            ctx = _build_row_context(
                ev.row,
                joined_by_row.get(id(ev.row)),
                join,
                sources,
                active_source_name,
                primary,
                data_index + 1,
                config_values,
                inputs,
                rows,
            )
            for tc in plan.cells:
                value = _render_cell(tc, ctx)
                style = style_cache.get((plan.template_row, tc.col))
                if tc.template.is_single_expression and style is not None:
                    value = _apply_numfmt_coercion(value, style[4])
                target = _write_cell_value(ws, out_row, tc.col, value)
                _apply_style(target, style)
            out_row += 1
            data_index += 1
            continue

        level_idx = ev.level - 1
        if level_idx >= len(plan.subtotal_rows):
            continue
        srow = plan.subtotal_rows[level_idx]
        subtotal_by_col = {sc.col: sc for sc in srow.subtotals}
        ctx = EvalContext(
            active_row={},
            active_source_name=active_source_name,
            active_source_columns=set(primary.headers) if primary.headers else None,
            inputs=inputs,
            config_values=config_values,
            active_row_set=ev.group_rows,
            named_sources=_build_named_sources_view(sources),
        )
        for tc in srow.cells:
            subtotal = subtotal_by_col.get(tc.col)
            if subtotal is not None:
                value = _eval_subtotal(subtotal, ev.group_rows)
            else:
                value = _render_cell(tc, ctx)
            style = style_cache.get((srow.template_row, tc.col))
            if (subtotal is not None or tc.template.is_single_expression) and style is not None:
                value = _apply_numfmt_coercion(value, style[4])
            target = _write_cell_value(ws, out_row, tc.col, value)
            _apply_style(target, style)
        out_row += 1
    return out_row


def _emit_vertical(
    ws: Any,
    plan: DataRowPlan,
    out_row: int,
    rows: list[dict[str, Any]],
    joined_rows: list[dict[str, Any] | None],
    join: JoinDirective | None,
    sources: dict[str, SourceData],
    active_source_name: str,
    primary: SourceData,
    style_cache: dict[tuple[int, int], Any],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
) -> int:
    for i, src_row in enumerate(rows):
        ctx = _build_row_context(
            src_row,
            joined_rows[i],
            join,
            sources,
            active_source_name,
            primary,
            i + 1,
            config_values,
            inputs,
            rows,
        )
        for tc in plan.cells:
            value = _render_cell(tc, ctx)
            style = style_cache.get((plan.template_row, tc.col))
            if tc.template.is_single_expression and style is not None:
                value = _apply_numfmt_coercion(value, style[4])
            target = _write_cell_value(ws, out_row, tc.col, value)
            _apply_style(target, style)
        out_row += 1
    return out_row


def _emit_repeat_right(
    ws: Any,
    plan: DataRowPlan,
    out_row: int,
    rows: list[dict[str, Any]],
    joined_rows: list[dict[str, Any] | None],
    join: JoinDirective | None,
    sources: dict[str, SourceData],
    active_source_name: str,
    primary: SourceData,
    style_cache: dict[tuple[int, int], Any],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
    col_span: int,
) -> int:
    if not rows:
        return out_row + 1
    base_col = min(tc.col for tc in plan.cells)
    for i, src_row in enumerate(rows):
        ctx = _build_row_context(
            src_row,
            joined_rows[i],
            join,
            sources,
            active_source_name,
            primary,
            i + 1,
            config_values,
            inputs,
            rows,
        )
        col_offset = i * col_span
        for tc in plan.cells:
            value = _render_cell(tc, ctx)
            style = style_cache.get((plan.template_row, tc.col))
            if tc.template.is_single_expression and style is not None:
                value = _apply_numfmt_coercion(value, style[4])
            new_col = tc.col + col_offset
            # First record reuses the original cell column; subsequent records
            # shift by `col_span` per record.
            target = _write_cell_value(ws, out_row, new_col, value)
            _apply_style(target, style)
            _ = base_col  # currently unused; kept for future left-anchor needs
    return out_row + 1


def _build_row_context(
    src_row: dict[str, Any],
    joined_row: dict[str, Any] | None,
    join: JoinDirective | None,
    sources: dict[str, SourceData],
    active_source_name: str,
    primary: SourceData,
    row_index: int,
    config_values: dict[str, Any],
    inputs: dict[str, Any],
    active_row_set: list[dict[str, Any]] | None,
) -> EvalContext:
    joined_rows: dict[str, dict[str, Any]] = {}
    joined_columns: dict[str, set[str]] = {}
    if join is not None and joined_row is not None:
        joined_rows[join.joined_source] = joined_row
        joined_columns[join.joined_source] = set(sources[join.joined_source].headers)
    return EvalContext(
        active_row=src_row,
        active_source_name=active_source_name,
        active_source_columns=set(primary.headers) if primary.headers else None,
        joined_rows=joined_rows,
        joined_columns=joined_columns,
        inputs=inputs,
        config_values=config_values,
        active_row_set=active_row_set,
        named_sources=_build_named_sources_view(sources),
        row_index=row_index,
    )


def _build_named_sources_view(sources: dict[str, SourceData]) -> dict[str, dict[str, Any]]:
    return {name: {"headers": sd.headers, "rows": sd.rows} for name, sd in sources.items()}


# ---------------------------------------------------------------------------
# numFmt-driven single-expression coercion (ADR-0003)
# ---------------------------------------------------------------------------


def _write_cell_value(ws, row, col, value):
    from openpyxl.cell.cell import TYPE_ERROR

    from .value_model import is_hyperlink_marker, is_xtl_error_cell

    if is_hyperlink_marker(value):
        cell = ws.cell(
            row=row,
            column=col,
            value=value.get("text") or value.get("__xl3_hyperlink__"),
        )
        cell.hyperlink = value["__xl3_hyperlink__"]
        return cell
    if is_xtl_error_cell(value):
        # ADR-0025: write a real Excel error cell (`#DIV/0!` etc.).
        cell = ws.cell(row=row, column=col, value=value["__xl3_error__"])
        cell.data_type = TYPE_ERROR
        return cell
    return ws.cell(row=row, column=col, value=value)


def _apply_numfmt_coercion(value: Any, number_format: str | None) -> Any:
    """ADR-0003: single-expression cells whose template cell has a date /
    number / text format MUST coerce the value to that format. Failures
    raise xl3/cell/numfmt-coercion.
    """
    if value is None or number_format is None:
        return value
    from .value_model import is_hyperlink_marker, is_xtl_error_cell

    if is_hyperlink_marker(value) or is_xtl_error_cell(value):
        return value
    nf = number_format
    if nf == "General":
        return value
    # Text format
    if nf == "@":
        if isinstance(value, str):
            return value
        return canonical_string(value)
    nf_lower = nf.lower()
    has_date_token = any(t in nf_lower for t in ("yyyy", "yy", "mm", "dd", "hh", "ss"))
    if has_date_token and not _is_pure_number_format(nf):
        return _coerce_to_date_for_numfmt(value, nf)
    if any(c in nf for c in "0#"):
        return _coerce_to_number_for_numfmt(value, nf)
    return value


def _is_pure_number_format(nf: str) -> bool:
    """Heuristic: format like `0`, `#,##0`, `0.00` is pure numeric, no date."""
    nf_lower = nf.lower()
    if any(t in nf_lower for t in ("yyyy", "yy", "mm", "dd", "hh", "ss")):
        return False
    return any(c in nf for c in "0#")


def _coerce_to_number_for_numfmt(value: Any, nf: str) -> Any:
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        from .value_model import parse_number_strict

        n = parse_number_strict(value)
        if n is None:
            raise xtl_error(
                "xl3/cell/numfmt-coercion",
                f'Value cannot be coerced to a number for cell format "{nf}": {value}',
            )
        return n
    raise xtl_error(
        "xl3/cell/numfmt-coercion",
        f'Value cannot be coerced to a number for cell format "{nf}": {canonical_string(value)}',
    )


def _coerce_to_date_for_numfmt(value: Any, nf: str) -> Any:
    from datetime import date, datetime

    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        s = value.strip()
        try:
            return datetime.fromisoformat(s)
        except ValueError as exc:
            raise xtl_error(
                "xl3/cell/numfmt-coercion",
                f'Value cannot be coerced to a date for cell format "{nf}": {value}',
            ) from exc
    raise xtl_error(
        "xl3/cell/numfmt-coercion",
        f'Value cannot be coerced to a date for cell format "{nf}": {canonical_string(value)}',
    )


def _render_cell(tc: TemplateCell, ctx: EvalContext) -> Any:
    tpl = tc.template
    if tpl.is_pure_text:
        # xl3 0.8.1 sync: pure-text cells re-emit the template's NATIVE
        # value (number/date/boolean) verbatim — re-rendering from the
        # cell's text form would stringify it (5500 → "5500").
        if tc.native_value is not None:
            return tc.native_value
        seg = tpl.segments[0]
        assert isinstance(seg, TextSegment)
        return seg.text if seg.text != "" else None

    if tpl.is_single_expression:
        seg = tpl.segments[0]
        assert isinstance(seg, ExprSegment)
        return evaluate(seg.expr, ctx)

    out: list[str] = []
    for seg in tpl.segments:
        if isinstance(seg, TextSegment):
            out.append(seg.text)
        elif isinstance(seg, ExprSegment):
            out.append(canonical_string(evaluate(seg.expr, ctx)))
        elif isinstance(seg, DirectiveSegment):
            # Directive cells should have been classified as directive rows
            # in parser.py and thus stripped before render. Reaching here
            # means a mixed cell — emit nothing for the directive segment.
            out.append("")
    return "".join(out)


def _evaluate_filename(
    parsed: ParsedTemplate,
    sources: dict[str, SourceData],
    config_values: dict[str, Any],
    inputs: dict[str, Any],
) -> str:
    pattern = parsed.meta.output_file_pattern or "output.xlsx"
    tpl = _parse_pattern(pattern)
    if tpl.is_pure_text:
        seg = tpl.segments[0]
        return seg.text if isinstance(seg, TextSegment) else pattern
    default_source = sources.get("default")
    row = default_source.rows[0] if default_source and default_source.rows else {}
    headers = set(default_source.headers) if default_source and default_source.headers else None
    ctx = EvalContext(
        active_row=row,
        inputs=inputs,
        config_values=config_values,
        active_source_columns=headers,
    )
    from .expression import BracketRef as _BracketRef

    out: list[str] = []
    for seg in tpl.segments:
        if isinstance(seg, TextSegment):
            out.append(seg.text)
        elif isinstance(seg, ExprSegment):
            value = evaluate(seg.expr, ctx)
            # ADR-0026: a bare-identifier `{{ Col }}` in a filename pattern
            # is a group-key reference; empty values substitute the
            # "(blank)" placeholder. `__config__[X]` / `__inputs__[X]` /
            # `Source[Col]` are NOT group keys and use canonical_string.
            if isinstance(seg.expr, _BracketRef):
                out.append(group_key_canonical(value))
            else:
                out.append(canonical_string(value))
    return "".join(out)


def _parse_pattern(pattern: str) -> CellTemplate:
    """Parse a filename pattern. Bare identifiers like `{{ Customer }}` are
    treated as group-key references per language.md §"Group Keys".
    """
    # Reuse the `{{ ... }}` splitter but parse each body with the relaxed
    # filename grammar.
    import re as _re

    from .expression import (
        DirectiveSegment as _DS,
    )
    from .expression import (
        ExprSegment as _ES,
    )
    from .expression import (
        TextSegment as _TS,
    )

    segs: list[Any] = []
    i = 0
    for m in _re.finditer(r"\{\{(.*?)\}\}", pattern, flags=_re.DOTALL):
        if m.start() > i:
            segs.append(_TS(pattern[i : m.start()]))
        body = m.group(1)
        if body.lstrip().startswith("@"):
            segs.append(_DS(body=body))
        else:
            from .expression import parse_filename_or_sheet_expression as _pf

            segs.append(_ES(expr=_pf(body), body=body))
        i = m.end()
    if i < len(pattern):
        segs.append(_TS(pattern[i:]))
    if not segs:
        segs.append(_TS(""))
    return CellTemplate(segments=segs)
