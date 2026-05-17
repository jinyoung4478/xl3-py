"""Source workbook reader.

Reads a data `.xlsx` according to `source_sheet` and `source_table` selectors
defined in `__config__`. Returns a list of row dicts (column → value).

Supports:
- `source_sheet`: literal name; prefix patterns ending in `*`
- `source_table = N`        (1-based header row, all rows below)
- `source_table = A1:D`     (open-ended range — header row 1, cols A..D)
- `source_table = A1:D200`  (finite range — rows 2..200)
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils import get_column_letter

from .errors import xtl_error
from .value_model import is_empty


@dataclass
class SourceData:
    sheet_name: str
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)


def _merge_master_for(
    ws: Any, row: int, col: int
) -> tuple[int, int] | None:
    """If (row, col) sits inside a merged range, return the master's (row, col).
    Returns None if not in any merge. ADR-0033/0035: porter MUST identify
    merges from merge-region metadata, not from cell-value presence —
    openpyxl returns None on slaves, ExcelJS returns the master's value.
    """
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return (mr.min_row, mr.min_col)
    return None


def _is_horizontal_merge_slave(
    ws: Any, row: int, col: int
) -> bool:
    """A cell is a horizontal-merge slave of the active header column if it
    sits inside a merged range whose master column is *to the left*. Per
    ADR-0033, such cells are transparent: they contribute no column and do
    not trigger duplicate-name."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.min_col != col
    return False


def _cell_value(cell: Cell) -> Any:
    """Extract a cell's source value per ADR-0017.

    Excel error cells (returned by openpyxl as strings starting with `#`) map
    to empty (None). Formula cells with cached results are handled by
    `data_only=True` at workbook load.
    """
    v = cell.value
    if v is None:
        return None
    if isinstance(v, CellRichText):
        return "".join(str(part) for part in v)
    # openpyxl returns Excel errors as strings like "#N/A", "#VALUE!", etc.
    if isinstance(v, str) and len(v) >= 2 and v[0] == "#" and v[-1] in "!?A0EFV":
        # Conservative match for the seven Excel error sentinels.
        if v in {"#N/A", "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NUM!", "#NULL!"}:
            return None
    return v


def _resolve_source_sheet(wb: Any, source_sheet: str | None) -> str:
    """ADR-0012 / evaluation.md §"Source Data Model"."""
    if source_sheet is None:
        return wb.sheetnames[0]
    # Exact match wins
    if source_sheet in wb.sheetnames:
        return source_sheet
    # Prefix pattern
    if source_sheet.endswith("*"):
        prefix = source_sheet[:-1]
        for name in wb.sheetnames:
            if name.startswith(prefix):
                return name
        raise xtl_error(
            "xl3/source/sheet-missing",
            f'Source sheet "{source_sheet}" was not found',
        )
    raise xtl_error(
        "xl3/source/sheet-missing",
        f'Source sheet "{source_sheet}" was not found',
    )


def _parse_source_table_row_shorthand(spec: str) -> int | None:
    """Parse `source_table = N` → 1-based header row. Returns None for ranges."""
    spec = spec.strip()
    try:
        n = int(spec)
    except ValueError:
        return None
    if n < 1:
        raise xtl_error(
            "xl3/config/invalid-source-table",
            f"source_table row numbers must be 1-based positive integers, got {n}",
        )
    return n


@dataclass
class _RangeSelector:
    """A `source_table = A1:D` or `A1:D200` selector."""

    left_col: int  # 1-based
    right_col: int  # 1-based
    header_row: int  # 1-based
    end_row: int | None  # 1-based, None for open-ended


_RANGE_RE = re.compile(
    r"^([A-Z]+)(\d+)\s*:\s*([A-Z]+)(\d+)?$",
)


def _parse_source_table_range(spec: str) -> _RangeSelector | None:
    """Parse `source_table = A1:D` (open) or `A1:D200` (finite). Returns
    None if `spec` isn't a range form."""
    from openpyxl.utils import column_index_from_string

    m = _RANGE_RE.match(spec.strip())
    if not m:
        return None
    left_col = column_index_from_string(m.group(1))
    header_row = int(m.group(2))
    right_col = column_index_from_string(m.group(3))
    end_row = int(m.group(4)) if m.group(4) else None
    if header_row < 1:
        raise xtl_error(
            "xl3/config/invalid-source-table",
            "source_table row numbers must be 1-based positive integers",
        )
    if left_col > right_col:
        raise xtl_error(
            "xl3/config/invalid-source-table",
            "source_table left column must not be right of the right column",
        )
    if end_row is not None and end_row < header_row:
        raise xtl_error(
            "xl3/config/invalid-source-table",
            "source_table end row must not be above the first row",
        )
    return _RangeSelector(
        left_col=left_col,
        right_col=right_col,
        header_row=header_row,
        end_row=end_row,
    )


def read_all_sources(
    source_bytes: bytes,
    default_sheet: str | None,
    default_table: str,
    declared_sources: list[Any],
) -> dict[str, "SourceData"]:
    """Read the default source plus every entry in `__sources__`.

    Returns a dict keyed by source name; the implicit default source is
    always present under the key `"default"` (per ADR-0012).
    """
    wb_values = load_workbook(BytesIO(source_bytes), data_only=True, rich_text=True)
    # Parallel load to detect formula cells that have NO cached result —
    # `data_only=True` masks them as `None`. We reconcile by checking
    # whether the same cell holds a `=...` formula in the formulas-only view.
    wb_formulas = load_workbook(BytesIO(source_bytes), data_only=False, rich_text=False)
    formula_view = _FormulaView(wb_formulas)
    out: dict[str, SourceData] = {}
    out["default"] = _read_one(wb_values, default_sheet, default_table, formula_view)
    for spec in declared_sources:
        out[spec.name] = _read_one(wb_values, spec.sheet, spec.table, formula_view)
    return out


@dataclass
class _FormulaView:
    """A way to peek at formula text per (sheet, row, col) without re-loading."""

    workbook: Any

    def is_uncached_formula(self, sheet: str, row: int, col: int) -> bool:
        if sheet not in self.workbook.sheetnames:
            return False
        c = self.workbook[sheet].cell(row=row, column=col)
        v = c.value
        return isinstance(v, str) and v.startswith("=")


def read_source(
    source_bytes: bytes,
    source_sheet: str | None,
    source_table: str = "1",
) -> SourceData:
    wb = load_workbook(BytesIO(source_bytes), data_only=True, rich_text=True)
    return _read_one(wb, source_sheet, source_table, formula_view=None)


def _read_one(
    wb: Any,
    source_sheet: str | None,
    source_table: str,
    formula_view: _FormulaView | None = None,
) -> SourceData:
    sheet_name = _resolve_source_sheet(wb, source_sheet)
    ws = wb[sheet_name]

    header_row_int = _parse_source_table_row_shorthand(source_table)
    if header_row_int is not None:
        return _read_with_inferred_span(ws, sheet_name, header_row_int, formula_view)
    rng = _parse_source_table_range(source_table)
    if rng is None:
        raise xtl_error(
            "xl3/config/invalid-source-table",
            f"source_table must be a positive integer or A1:D[N] range, got {source_table!r}",
        )
    return _read_with_explicit_range(ws, sheet_name, rng, formula_view)


def _read_with_inferred_span(
    ws: Any,
    sheet_name: str,
    header_row: int,
    formula_view: _FormulaView | None,
) -> SourceData:
    """Row-shorthand form: header row is N; column span = first..last
    non-empty header cell on that row. ADR-0033: a header that is a
    horizontal-merge slave inherits the master's value (master is to the
    left), so the span includes such cells. We detect non-emptiness via the
    master-anchored read, not via raw cell value."""
    if not ws.max_row or ws.max_row < header_row:
        return SourceData(sheet_name=sheet_name)
    first_idx: int | None = None
    last_idx: int | None = None
    for col_idx in range(1, ws.max_column + 1):
        text = _header_cell_text_with_merges(ws, header_row, col_idx)
        if text is not None and text.strip() != "":
            if first_idx is None:
                first_idx = col_idx
            last_idx = col_idx
    if first_idx is None:
        return SourceData(sheet_name=sheet_name)
    headers, col_indices = _read_header_row(
        ws, sheet_name, header_row, first_idx, last_idx, formula_view  # type: ignore[arg-type]
    )
    rows = _read_data_rows(
        ws,
        sheet_name,
        header_row,
        col_indices,
        headers,
        end_row=None,
        formula_view=formula_view,
    )
    return SourceData(sheet_name=sheet_name, headers=headers, rows=rows)


def _read_with_explicit_range(
    ws: Any,
    sheet_name: str,
    rng: "_RangeSelector",
    formula_view: _FormulaView | None,
) -> SourceData:
    headers, col_indices = _read_header_row(
        ws, sheet_name, rng.header_row, rng.left_col, rng.right_col, formula_view
    )
    rows = _read_data_rows(
        ws,
        sheet_name,
        rng.header_row,
        col_indices,
        headers,
        end_row=rng.end_row,
        formula_view=formula_view,
    )
    return SourceData(sheet_name=sheet_name, headers=headers, rows=rows)


def _read_header_row(
    ws: Any,
    sheet_name: str,
    header_row: int,
    left_col: int,
    right_col: int,
    formula_view: _FormulaView | None,
) -> tuple[list[str], list[int]]:
    """ADR-0017 effective text + ADR-0033 merged-header transparency.

    Iterates `left_col..right_col`. For each column index:
      - If the cell is a horizontal-merge slave (master in a different
        column inside the header band), skip it. It contributes neither
        a header name nor a data column.
      - Otherwise, read the master-anchored value (handles vertical and
        2D merges where the header row falls on a slave row but the
        master is in an earlier row of the same column).

    Returns (headers, col_indices) so the data-row reader knows which
    physical columns to read for each logical column.
    """
    headers: list[str] = []
    col_indices: list[int] = []
    seen: set[str] = set()
    for col_idx in range(left_col, right_col + 1):
        if _is_horizontal_merge_slave(ws, header_row, col_idx):
            continue
        v = _header_cell_text_with_merges(ws, header_row, col_idx)
        if (v is None or v.strip() == "") and formula_view is not None:
            master = _merge_master_for(ws, header_row, col_idx)
            mrow, mcol = master if master is not None else (header_row, col_idx)
            if formula_view.is_uncached_formula(sheet_name, mrow, mcol):
                raise xtl_error(
                    "xl3/cell/formula-no-cache",
                    f"Formula cell {get_column_letter(mcol)}{mrow} has no cached result",
                )
        if v is None or v.strip() == "":
            raise xtl_error(
                "xl3/source/missing-header",
                f"source_table header cell {get_column_letter(col_idx)}{header_row} is empty (merged header band: no master in window)",
            )
        name = v.strip()
        if name in seen:
            raise xtl_error(
                "xl3/source/duplicate-name",
                f'source_table has duplicate header "{name}"',
            )
        seen.add(name)
        headers.append(name)
        col_indices.append(col_idx)
    return headers, col_indices


def _header_cell_text_with_merges(ws: Any, row: int, col: int) -> str | None:
    """Read header text from a cell, dereferencing merges (ADR-0033)."""
    master = _merge_master_for(ws, row, col)
    if master is not None:
        mrow, mcol = master
        cell = ws.cell(row=mrow, column=mcol)
    else:
        cell = ws.cell(row=row, column=col)
    return _header_cell_text(cell)


def _header_cell_text(cell: Cell) -> str | None:
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, CellRichText):
        return "".join(str(part) for part in v)
    if isinstance(v, str) and v.startswith("="):
        raise xtl_error(
            "xl3/cell/formula-no-cache",
            f"Formula cell {cell.coordinate} has no cached result",
        )
    return str(v)


def _read_data_rows(
    ws: Any,
    sheet_name: str,
    header_row: int,
    col_indices: list[int],
    headers: list[str],
    end_row: int | None,
    formula_view: _FormulaView | None,
) -> list[dict[str, Any]]:
    """Iterate data rows starting at header_row+1. For each kept (master)
    column, read the master value of any merged range (ADR-0035 broadcast).

    Subtlety: if the header band spans multiple rows (2D merge), the
    first data row is the row immediately AFTER the entire header band
    ends. We compute that by finding the bottom of any merge whose top
    is the header_row.
    """
    rows: list[dict[str, Any]] = []
    first_data_row = header_row + 1
    for mr in ws.merged_cells.ranges:
        if mr.min_row == header_row and mr.max_row >= first_data_row:
            first_data_row = max(first_data_row, mr.max_row + 1)
    last_row = end_row if end_row is not None else (ws.max_row or header_row)
    for r in range(first_data_row, last_row + 1):
        row_dict: dict[str, Any] = {}
        row_empty = True
        for col_name, col_idx in zip(headers, col_indices, strict=True):
            master = _merge_master_for(ws, r, col_idx)
            if master is not None:
                mrow, mcol = master
                cell = ws.cell(row=mrow, column=mcol)
                read_row, read_col = mrow, mcol
            else:
                cell = ws.cell(row=r, column=col_idx)
                read_row, read_col = r, col_idx
            val = _cell_value(cell)
            if val is None and formula_view is not None:
                if formula_view.is_uncached_formula(sheet_name, read_row, read_col):
                    raise xtl_error(
                        "xl3/cell/formula-no-cache",
                        f"Formula cell {cell.coordinate} has no cached result",
                    )
            row_dict[col_name] = val
            if not is_empty(val):
                row_empty = False
        if not row_empty:
            rows.append(row_dict)
    return rows
